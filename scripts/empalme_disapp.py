#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Empalme Disapp -> Presta Ya. Importador idempotente y auditable.
Ver docs/mapeo-empalme.md para el mapeo y las decisiones.

Modos:
  --audit                 Solo lectura: consolida los .xlsx, reporta cobertura de
                          fechas, dias faltantes, vendedores, huerfanos y el GAP de
                          reconstruccion de pagos no-diarios. NO toca DB.
  --probe REFS            Cuadre dirigido (coma-sep) sin escribir: muestra Saldo/Cuotas
                          de Disapp vs el carton tras reconstruir. Ej:
                          --probe PRD0003208827,PRD0002535750
  --dry-run  (default)    Consolida + reporta que HARIA (cuentas, banderas, reconstruccion),
                          sin escribir. Emite reconstruccion_revision.csv.
  --commit                Escribe de verdad (crea cobradores/auth, clientes, creditos,
                          stubs, pagos y pagos de AJUSTE). Requiere --env-file.
  --prod                  Permite apuntar a produccion (si no, aborta cuando el .env no
                          parece de prueba). NUNCA usar sin intencion.
  --validate-sample N     Tras --commit: toma N clientes y muestra saldo/estado calculado.
  --env-file PATH         .env con SUPABASE_URL/NEXT_PUBLIC_SUPABASE_URL + SERVICE_ROLE_KEY
                          (default .env.prueba).
  --src PATH              Carpeta de exports (default C:\\Users\\Carlos\\migracion).
  --out PATH              Carpeta de reportes (default <src>\\_reportes).

Reglas de dinero (criticas):
  · creditos_*.xlsx: montos como TEXTO uruguayo '6.000,00' -> 6000.00
  · recaudos_*.xlsx: Excel se comio el separador de miles. float -> x1000, int -> tal cual.
  · Fechas dd/mm/yyyy. Los pagos NUNCA se borran; 'Saldo Pendiente' es metadato, no fuente.
  · El carton es FIFO ACUMULADO (lib/cartones.ts): el estado de cada cuota sale del
    TOTAL pagado, NO del dia_credito ("Cuota #" de Disapp es un snapshot inutil).
  · Reconstruccion (paso 6): Disapp solo exporta recaudos DIARIOS; los pagos de creditos
    semanales/quincenales/mensuales/VIP se reconstruyen desde la columna 'Pagos' del
    Excel de creditos, sembrando ajustes (origen='ajuste_migracion'). Verdad = el DINERO
    ('Pagos'+'Saldo'=='Total' en 100% de activos); 'Cuotas Pend.' es solo informativo
    (concuerda con el dinero en ~26%).
"""
import argparse, glob, os, sys, json, re, secrets, urllib.request, urllib.parse, urllib.error
import datetime as dt
from collections import defaultdict, Counter

# La consola de Windows (cp1252) no traga los símbolos → ✓ ⚠ de los prints. Forzar
# UTF-8 evita que el importador crashee a mitad del go-live por un carácter.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl:  python -m pip install openpyxl")

DEFAULT_SRC = r"C:\Users\Carlos\migracion"
MODALIDAD_FREC = {"diaria": "diario", "semanal": "semanal", "quincenal": "quincenal",
                  "mensual": "mensual", "personalizado": "diario"}
MESES = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]
BATCH = 500

# ══ Parsers ═════════════════════════════════════════════════════════════════
def parse_money_text(v):
    """Montos de creditos: '6.000,00' -> 6000.00. Tambien acepta numeros."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None

def parse_recaudo(v):
    """recaudos con el bug del separador de miles: int -> tal cual; float -> x1000;
       str -> parseo uruguayo."""
    if v is None or v == "" or isinstance(v, bool):
        return None
    if isinstance(v, int):
        return float(v)
    if isinstance(v, float):
        return round(v * 1000.0, 2)
    return parse_money_text(str(v))

def parse_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    m = re.match(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})", str(v))
    if m:
        d, mo, y = map(int, m.groups())
        try:
            return dt.date(y, mo, d)
        except ValueError:
            return None
    return None

def iso_ts(d):
    """date -> timestamptz al mediodia de Uruguay (evita corrimiento de dia por TZ)."""
    return f"{d.isoformat()}T12:00:00-03:00" if d else None

# ══ Lectura xlsx ════════════════════════════════════════════════════════════
def discover(src, prefix):
    return sorted(glob.glob(os.path.join(src, prefix + "*.xlsx")))

def header_map(ws):
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(c).strip().lower(): i for i, c in enumerate(row1) if c is not None}

def col(hm, *names):
    for n in names:
        i = hm.get(n.strip().lower())
        if i is not None:
            return i
    return None

def rows_of(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hm = header_map(ws)
    for r in ws.iter_rows(min_row=2, values_only=True):
        yield hm, r
    wb.close()

def _raw(r, i): return r[i] if i is not None else None
def _s(r, i):   return str(r[i]).strip() if (i is not None and r[i] not in (None, "")) else None
def _int(r, i):
    if i is None or r[i] in (None, ""):
        return None
    try:
        return int(float(str(r[i]).replace(",", ".")))
    except ValueError:
        return None

def es_supervisor(nombre): return bool(nombre) and "SUPERVISOR" in nombre.upper()
def es_admin_vend(nombre): return bool(nombre) and "ADMINISTRADOR" in nombre.upper()

# ══ Consolidacion (pura, sin DB) ════════════════════════════════════════════
def load_clientes(src):
    clientes = {}     # disapp_id -> {..}
    docs_vistos = {}  # documento -> disapp_id (el 1ro que lo uso)
    dup_docs = 0
    for path in discover(src, "clientes"):
        for hm, r in rows_of(path):
            ic = col(hm, "ID")
            if ic is None or r[ic] in (None, ""):
                continue
            did = str(r[ic]).strip()
            if did in clientes:
                continue
            doc = _s(r, col(hm, "Documento"))
            doc_final = doc
            if doc:
                if doc in docs_vistos:      # decision 2: 1ro conserva, dup -> NULL
                    doc_final = None
                    dup_docs += 1
                else:
                    docs_vistos[doc] = did
            clientes[did] = {
                "disapp_id": did,
                "nombre": _s(r, col(hm, "Nombre")) or "(sin nombre)",
                "documento": doc_final,
                "documento_original": doc,
                "telefono": _s(r, col(hm, "Teléfono", "Telefono")),
                "direccion": _s(r, col(hm, "Dirección", "Direccion")),
                "notas": _s(r, col(hm, "Observación", "Observacion")),
                "vendedor": _s(r, col(hm, "Vendedor")),
                "activo": (_s(r, col(hm, "Estado")) or "").lower() != "inactivo",
            }
    return clientes, docs_vistos, dup_docs

def load_creditos(src):
    by_id = {}
    vendedores = {}   # id_vendedor -> nombre
    vend_por_texto = {}  # nombre texto -> id_vendedor
    for path in discover(src, "creditos"):
        for hm, r in rows_of(path):
            idc = col(hm, "ID Crédito", "ID Credito")
            if idc is None or r[idc] in (None, ""):
                continue
            idv = _s(r, col(hm, "ID Vendedor"))
            vnom = _s(r, col(hm, "Vendedor"))
            if idv and vnom:
                vendedores[idv] = vnom
                vend_por_texto[vnom] = idv
            mod = (_s(r, col(hm, "Modalidad")) or "diaria").lower()
            by_id[str(r[idc]).strip()] = {
                "disapp_credit_id": str(r[idc]).strip(),
                "ref": _s(r, col(hm, "Crédito #", "Credito #")),
                "id_cliente": _s(r, col(hm, "ID Cliente")),
                "id_vendedor": idv,
                "vendedor": vnom,
                "monto_prestado": parse_money_text(_raw(r, col(hm, "Valor Crédito", "Valor Credito"))),
                "cuota": parse_money_text(_raw(r, col(hm, "Valor Cuota"))),
                "cuotas": _int(r, col(hm, "Cuotas")),
                "frecuencia": MODALIDAD_FREC.get(mod, "diario"),
                "modalidad_rara": mod not in MODALIDAD_FREC or mod == "personalizado",
                "fecha": parse_date(_raw(r, col(hm, "Fecha Crédito", "Fecha Credito"))),
                # Columnas de reconstrucción de pagos no-diarios (verificadas en el
                # Excel real). Disapp NO exporta los pagos semanales/quincenales/
                # mensuales/VIP → esas columnas son la ÚNICA fuente de lo pagado.
                # Nota crítica (verificada en data real 2026-07-08): 'Pagos'+'Saldo
                # Pendiente'=='Total c/ Intereses' en el 100% de los activos, pero
                # 'Cuotas Pend.' solo concuerda con el dinero en el 25,6% → se usa
                # el DINERO como verdad, 'Cuotas Pend.' es solo informativo.
                "pagos_disapp": parse_money_text(_raw(r, col(hm, "Pagos"))),
                "saldo_pendiente": parse_money_text(_raw(r, col(hm, "Saldo Pendiente"))),
                "total_c_intereses": parse_money_text(_raw(r, col(hm, "Total c/ Intereses", "Total c/ Interes", "Total con Intereses"))),
                "cuotas_pend": _int(r, col(hm, "Cuotas Pend.", "Cuotas Pend")),
                "estado_disapp": (_s(r, col(hm, "Estado")) or "").strip(),
            }
    return by_id, vendedores, vend_por_texto

def load_pagos(src):
    pagos = {}
    files = discover(src, "recaudos")
    for path in files:
        for hm, r in rows_of(path):
            ip = col(hm, "ID Pago")
            if ip is None or r[ip] in (None, ""):
                continue
            key = str(r[ip]).strip()
            if key in pagos:
                continue
            raw_rec = _raw(r, col(hm, "Recaudo"))
            pagos[key] = {
                "id_pago": key,
                "ref": _s(r, col(hm, "Ref. Crédito", "Ref. Credito")),
                "vendedor": _s(r, col(hm, "Vendedor")),
                "documento": _s(r, col(hm, "Documento")),
                "monto": parse_recaudo(raw_rec),
                "corregido": isinstance(raw_rec, float),
                "cuota_num": _int(r, col(hm, "Cuota #")),
                "total_cuotas": _int(r, col(hm, "Total Cuotas")),
                "valor_cuota": parse_recaudo(_raw(r, col(hm, "Valor Cuota"))),
                "total_credito": parse_recaudo(_raw(r, col(hm, "Total Crédito", "Total Credito"))),
                "fecha": parse_date(_raw(r, col(hm, "Fecha Pago"))),
            }
    return pagos, files

def consolidar(src):
    clientes, docs, dup_docs = load_clientes(src)
    creditos, vendedores, vend_por_texto = load_creditos(src)
    pagos, prec_files = load_pagos(src)
    # refs de creditos activos (por "Credito #")
    refs_activos = {c["ref"]: cid for cid, c in creditos.items() if c["ref"]}
    # huerfanos: refs en pagos que no son de un credito activo
    huerfanos = defaultdict(list)   # ref -> [pago,..]
    for p in pagos.values():
        if p["ref"] and p["ref"] not in refs_activos:
            huerfanos[p["ref"]].append(p)
    return {
        "clientes": clientes, "docs": docs, "dup_docs": dup_docs,
        "creditos": creditos, "vendedores": vendedores, "vend_por_texto": vend_por_texto,
        "pagos": pagos, "prec_files": prec_files,
        "refs_activos": refs_activos, "huerfanos": huerfanos,
    }

# ══ RECONSTRUCCIÓN de pagos no-diarios ══════════════════════════════════════
# Disapp solo exporta los recaudos de créditos DIARIOS. Los pagos de créditos
# semanales / quincenales / mensuales / VIP de supervisor NO se pueden exportar
# → entran con $0 pagado e inflan mora y cartera. La única fuente de lo pagado
# para esos créditos es la columna 'Pagos' del Excel de créditos.
#
# ACOPLE con el cartón (lib/cartones.ts): el cartón es FIFO ACUMULADO — el estado
# de cada cuota se deriva del TOTAL pagado (Math.min(cuota, totalPagado - i*cuota)),
# NO del dia_credito de cada pago (el 'Cuota #' de Disapp es un snapshot inútil).
# ⇒ para que mora y saldo salgan exactos basta con UNA identidad:
#       suma(pagos del crédito) == Pagos (columna Disapp)
# La distribución cuota-por-cuota de abajo es cosmética (comprobante natural +
# respeta el trigger dia_credito ≤ total_dias); el resultado en dinero es idéntico
# a un solo lump. Se distribuye llenando las cuotas MÁS TEMPRANAS aún no cubiertas
# por recaudos reales, con cobertura FIFO-consistente (no por el 'Cuota #' basura).
TOL_MONTO = 1.0

def _cubierto_fifo(cubierto, dia, cuota):
    """Cuánto de la cuota `dia` (1-based) queda cubierto por `cubierto` acumulado,
    bajo FIFO (igual que el cartón). Se satura a [0, cuota]."""
    return max(0.0, min(cuota, cubierto - (dia - 1) * cuota))

def reconstruir_creditos(d, only_refs=None):
    """PURA (no toca DB). Para cada crédito ACTIVO cuya columna 'Pagos' supera lo
    ya importado de recaudos reales, calcula los pagos de AJUSTE a sembrar.
    Devuelve dict ref -> {credit_id, fecha, cuota, total_dias, ya, pagos_disapp,
    reconstruir, sembrados:[(dia,monto)], total_pagado_final, falta_final,
    saldo_pend, cuotas_pend, revision:[str]}.
    Solo créditos con Estado=Activo (los que definen el cartón de go-live)."""
    # Suma de recaudos REALES ya imputados por crédito activo (por ref = 'Crédito #').
    ya_por_ref = defaultdict(float)
    for p in d["pagos"].values():
        if p["ref"] and p["ref"] in d["refs_activos"] and p["monto"]:
            ya_por_ref[p["ref"]] += p["monto"]

    res = {}
    for cid, c in d["creditos"].items():
        ref = c["ref"]
        if not ref:
            continue
        if (c.get("estado_disapp") or "").lower() != "activo":
            continue
        if only_refs and ref not in only_refs:
            continue

        cuota = c["cuota"] or 0.0
        total_dias = c["cuotas"] or 0
        pagos_disapp = c["pagos_disapp"] or 0.0
        total_ci = c["total_c_intereses"]
        saldo_pend = c["saldo_pendiente"]
        cuotas_pend = c["cuotas_pend"]
        ya = round(ya_por_ref.get(ref, 0.0), 2)
        rev = []

        # Precheck de consistencia: el cartón usa totalAPagar = cuota*total_dias.
        # Si eso no coincide con 'Total c/ Intereses', el saldo reconstruido no va
        # a igualar 'Saldo Pendiente' aunque los pagos estén perfectos (INFORMATIVO).
        if cuota and total_dias and total_ci is not None:
            if abs(cuota * total_dias - total_ci) > TOL_MONTO:
                rev.append(f"cuota*cuotas={cuota*total_dias:.0f} != Totalc/Int={total_ci:.0f}")

        info = {
            "credit_id": cid, "ref": ref, "fecha": c["fecha"],
            "cuota": cuota, "total_dias": total_dias, "ya": ya,
            "pagos_disapp": pagos_disapp, "saldo_pend": saldo_pend,
            "cuotas_pend": cuotas_pend, "frecuencia": c["frecuencia"],
            "sembrados": [], "reconstruir": 0.0,
            "total_pagado_final": ya, "falta_final": None, "revision": rev,
        }

        if cuota <= 0 or total_dias <= 0:
            rev.append("cuota/total_dias inválidos")
            info["falta_final"] = max(0.0, cuota * total_dias - ya)
            res[ref] = info
            continue

        total_carton = cuota * total_dias
        reconstruir = round(pagos_disapp - ya, 2)

        # Cruce k por monto vs k por cuotas (INFORMATIVO: 'Cuotas Pend.' solo
        # concuerda con el dinero en ~25% de los casos → nunca bloquea).
        if cuotas_pend is not None:
            k_monto = pagos_disapp / cuota
            k_cuotas = total_dias - cuotas_pend
            if abs(round(k_monto) - k_cuotas) > 1:
                rev.append(f"k_monto={k_monto:.1f} vs k_cuotas={k_cuotas} (Cuotas Pend. no cuadra)")

        if reconstruir < 0:
            # Los recaudos reales traen MÁS que el Excel → no sembrar; a revisión.
            rev.append(f"ya_importado={ya:.0f} > Pagos={pagos_disapp:.0f} (no siembra)")
        elif reconstruir > 0:
            # Cap para no exceder el cartón (sobrepago / dato inconsistente).
            espacio = round(total_carton - ya, 2)
            if reconstruir > espacio + 0.5:
                rev.append(f"Pagos>cartón: reconstruir={reconstruir:.0f} > espacio={espacio:.0f} (cap)")
                reconstruir = max(0.0, espacio)
            restante = reconstruir
            for dia in range(1, total_dias + 1):
                if restante <= 0:
                    break
                falta_dia = round(cuota - _cubierto_fifo(ya, dia, cuota), 2)
                if falta_dia <= 0:
                    continue
                monto = round(min(falta_dia, restante), 2)
                if monto <= 0:
                    continue
                info["sembrados"].append((dia, monto))
                restante = round(restante - monto, 2)
            # Residuo de centavos: al último sembrado; nunca crear día > total_dias.
            if restante > 0.5:
                rev.append(f"sobró {restante:.2f} sin cuota donde ubicar")
            elif restante > 0 and info["sembrados"]:
                d0, m0 = info["sembrados"][-1]
                info["sembrados"][-1] = (d0, round(m0 + restante, 2))
            info["reconstruir"] = round(sum(m for _, m in info["sembrados"]), 2)

        info["total_pagado_final"] = round(ya + info["reconstruir"], 2)
        info["falta_final"] = round(max(0.0, total_carton - info["total_pagado_final"]), 2)
        res[ref] = info
    return res

def _reconciliacion_identidades(info):
    """Devuelve la lista de identidades que FALLAN para un crédito reconstruido
    (contra las columnas del Excel). #1 y #2 son duras (dinero); #3 informativa."""
    fallos = []
    cuota, td = info["cuota"], info["total_dias"]
    tp = info["total_pagado_final"]
    # #1 suma de pagos == Pagos
    if info["pagos_disapp"] is not None and abs(tp - info["pagos_disapp"]) > TOL_MONTO:
        fallos.append(("1_suma!=Pagos", info["pagos_disapp"], tp))
    # #2 falta del cartón == Saldo Pendiente
    if info["saldo_pend"] is not None and abs(info["falta_final"] - info["saldo_pend"]) > TOL_MONTO:
        fallos.append(("2_falta!=Saldo", info["saldo_pend"], info["falta_final"]))
    # #3 cuotas pagadas completas == Cuotas - Cuotas Pend. (INFORMATIVA)
    if info["cuotas_pend"] is not None and cuota > 0:
        pagadas = int(tp // cuota) if td else 0
        esperado = td - info["cuotas_pend"]
        if abs(pagadas - esperado) > 1:
            fallos.append(("3_cuotas(info)", esperado, pagadas))
    return fallos

def _filas_ajuste(res, ref_to_uuid, ahora_iso):
    """Construye las filas de pago de ajuste para insertar. Solo para créditos ya
    presentes en `ref_to_uuid` (prestamos insertados). Idempotente por
    disapp_pago_id = 'ajuste-<credit_id>-<dia>'."""
    filas = []
    for ref, info in res.items():
        pid = ref_to_uuid.get(ref)
        if not pid or not info["sembrados"]:
            continue
        # registrado_en al inicio del crédito → NO contamina "recaudado hoy/mes".
        reg = iso_ts(info["fecha"]) or iso_ts(dt.date.today())
        for dia, monto in info["sembrados"]:
            filas.append({
                "prestamo_id": pid, "dia_credito": dia, "monto": monto,
                "registrado_por": None, "registrado_en": reg,
                "origen": "ajuste_migracion", "importado_en": ahora_iso,
                "disapp_pago_id": f"ajuste-{info['credit_id']}-{dia}",
                "disapp_credit_ref": ref,
            })
    return filas

# ══ AUDIT ═══════════════════════════════════════════════════════════════════
def audit(src, out):
    print("Leyendo exports de:", src)
    d = consolidar(src)
    clientes, creditos, pagos = d["clientes"], d["creditos"], d["pagos"]
    corregidos = sum(1 for p in pagos.values() if p["corregido"])
    total = sum(p["monto"] for p in pagos.values() if p["monto"])
    print("\n" + "=" * 68 + "\nRESUMEN\n" + "=" * 68)
    print(f"  Clientes unicos:                 {len(clientes):>8}")
    print(f"  Creditos activos:                {len(creditos):>8}")
    print(f"  Archivos de recaudos:            {len(d['prec_files']):>8}")
    print(f"  Pagos unicos:                    {len(pagos):>8}")
    print(f"  Pagos corregidos x1000:          {corregidos:>8}")
    print(f"  Documentos duplicados (a NULL):  {d['dup_docs']:>8}")
    print(f"  Suma recaudo (corregida):      $ {total:>14,.0f}")

    por_mes, suma_mes = Counter(), defaultdict(float)
    fechas = []
    for p in pagos.values():
        if p["fecha"]:
            fechas.append(p["fecha"]); mk = (p["fecha"].year, p["fecha"].month)
            por_mes[mk] += 1; suma_mes[mk] += p["monto"] or 0
    if fechas:
        print(f"\n  Rango: {min(fechas)}  ->  {max(fechas)}")
        print("  --- Pagos y $ por mes ---")
        for mk in sorted(por_mes):
            y, m = mk
            print(f"     {MESES[m-1]}-{y}:  {por_mes[mk]:>6} pagos   $ {suma_mes[mk]:>14,.0f}")

    # Gap de reconstrucción: créditos activos donde Disapp dice que pagaron MÁS
    # de lo que hay en recaudos (los no-diarios). Sin escribir nada.
    res = reconstruir_creditos(d)
    recon = [i for i in res.values() if i["sembrados"]]
    gap = round(sum(i["reconstruir"] for i in recon), 2)
    por_frec = Counter(i["frecuencia"] for i in recon)
    print(f"\n  --- Gap de pagos no-diarios (reconstrucción) ---")
    print(f"  Créditos activos con gap:  {len(recon)}   monto a reconstruir: $ {gap:,.0f}")
    print(f"  Pagos de ajuste a sembrar: {sum(len(i['sembrados']) for i in recon)}")
    print(f"  Por frecuencia: {dict(por_frec)}")

    pagos_huerf = sum(len(v) for v in d["huerfanos"].values())
    print(f"\n  Refs huerfanas (stub): {len(d['huerfanos'])}   pagos afectados: {pagos_huerf}")
    sup = sorted({v for v in d["vendedores"].values() if es_supervisor(v)})
    print(f"  Vendedores: {len(d['vendedores'])}  (supervisores: {len(sup)})")
    for s in sup:
        print("    [SUP]", s)

    os.makedirs(out, exist_ok=True)
    with open(os.path.join(out, "audit.json"), "w", encoding="utf-8") as f:
        json.dump({"clientes": len(clientes), "creditos": len(creditos),
                   "pagos": len(pagos), "corregidos_x1000": corregidos,
                   "dup_docs": d["dup_docs"], "suma": round(total, 2),
                   "huerfanos": len(d["huerfanos"]), "pagos_huerfanos": pagos_huerf,
                   "vendedores": sorted(d["vendedores"].values())},
                  f, ensure_ascii=False, indent=2)
    print(f"\n  Reporte: {os.path.join(out, 'audit.json')}")

# ══ DB (PostgREST + GoTrue admin) ═══════════════════════════════════════════
def load_env(path):
    env = {}
    if os.path.isfile(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env

def http(url, method, key, body=None, prefer=None):
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("apikey", key); req.add_header("Authorization", "Bearer " + key)
    req.add_header("Content-Type", "application/json")
    if prefer:
        req.add_header("Prefer", prefer)
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            txt = resp.read().decode()
            return resp.status, (json.loads(txt) if txt else None)
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode()

def upsert(db, table, rows, on_conflict, ignore=False, rep=True):
    if not rows:
        return []
    res = "ignore-duplicates" if ignore else "merge-duplicates"
    prefer = f"resolution={res},return={'representation' if rep else 'minimal'}"
    base = db["url"] + "/rest/v1/" + table + "?on_conflict=" + on_conflict
    out = []
    for i in range(0, len(rows), BATCH):
        st, data = http(base, "POST", db["key"], rows[i:i+BATCH], prefer)
        if st >= 300:
            raise RuntimeError(f"upsert {table} [{st}]: {str(data)[:500]}")
        if rep and isinstance(data, list):
            out.extend(data)
        sys.stdout.write(f"\r    {table}: {min(i+BATCH,len(rows))}/{len(rows)}")
        sys.stdout.flush()
    print()
    return out

def get_rows(db, table, select, params=None):
    """GET paginado (para armar mapas grandes). ⚠️ SIEMPRE con order estable: sin
    él, PostgREST puede repetir/saltear filas entre páginas (>1000) y dejar mapas
    incompletos → créditos/pagos perdidos. Mismo bug de dinero que en el panel."""
    out = []
    off = 0
    while True:
        q = {"select": select, "limit": 1000, "offset": off, "order": "id.asc"}
        if params:
            q.update(params)
        url = db["url"] + "/rest/v1/" + table + "?" + urllib.parse.urlencode(q)
        st, data = http(url, "GET", db["key"])
        if st >= 300:
            raise RuntimeError(f"get {table} [{st}]: {str(data)[:300]}")
        out.extend(data or [])
        if not data or len(data) < 1000:
            break
        off += 1000
    return out

def crear_auth(db, email, password):
    st, data = http(db["url"] + "/auth/v1/admin/users", "POST", db["key"],
                    {"email": email, "password": password, "email_confirm": True})
    if st >= 300 or not isinstance(data, dict):
        return None, f"[{st}] {str(data)[:200]}"
    return data.get("id"), None

def rand_pass():
    abc = "abcdefghijkmnpqrstuvwxyzABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return "".join(secrets.choice(abc) for _ in range(12))

# ══ COMMIT ══════════════════════════════════════════════════════════════════
def commit_import(src, out, db, dry):
    d = consolidar(src)
    os.makedirs(out, exist_ok=True)
    modo = "DRY-RUN (no escribe)" if dry else "COMMIT (escribe)"
    destino = db["host"] if db else "sin conexion (conteos previos a la DB)"
    print(f"\n=== IMPORT {modo} → {destino} ===")

    # ── 1. Cobradores / supervisores (usuarios + auth) ─────────────────────
    admin_id = None
    if not dry:
        admins = get_rows(db, "usuarios", "id", {"rol": "eq.admin", "limit": 1})
        admin_id = admins[0]["id"] if admins else None
    id_vend_to_uuid = {}     # id_vendedor -> usuario uuid
    creds_csv = []
    print(f"\n[1] Cobradores/supervisores: {len(d['vendedores'])}")
    for idv, nombre in sorted(d["vendedores"].items()):
        # Los "supervisores" de Disapp (CESAR/BOSO/EDWIN) NO llevan ruta diaria:
        # gestionan una cartera VIP (semanal, 0%, montos altos) de clientes reales.
        # Se importan como un cobrador dedicado "CARTERA <zona>" para que esa
        # cartera quede cobrable en Presta Ya (decisión Carlos 2026-07-08).
        if es_supervisor(nombre):
            nombre = nombre.replace("SUPERVISOR", "CARTERA")
        rol = "cobrador"
        if es_admin_vend(nombre) and admin_id:
            id_vend_to_uuid[idv] = admin_id
            continue
        if dry:
            continue
        # idempotencia: ya existe por disapp_vendedor_id?
        ya = get_rows(db, "usuarios", "id", {"disapp_vendedor_id": f"eq.{idv}", "limit": 1})
        if ya:
            id_vend_to_uuid[idv] = ya[0]["id"]
            continue
        email = f"cobrador-{idv}@import.prestaya.local"
        pwd = rand_pass()
        auth_id, err = crear_auth(db, email, pwd)
        if err:
            print(f"    ! auth {nombre}: {err}"); continue
        fila = upsert(db, "usuarios", [{
            "nombre": nombre, "rol": rol, "activo": True,
            "auth_user_id": auth_id, "disapp_vendedor_id": idv,
        }], "disapp_vendedor_id")
        if fila:
            id_vend_to_uuid[idv] = fila[0]["id"]
            creds_csv.append(f"{rol},{nombre},{email},{pwd}")
    if creds_csv:
        p = os.path.join(out, "credenciales_cobradores.csv")
        with open(p, "w", encoding="utf-8") as f:
            f.write("rol,nombre,email,password\n" + "\n".join(creds_csv) + "\n")
        print(f"    Credenciales -> {p}  ({len(creds_csv)} nuevas)")
    vtexto_to_uuid = {t: id_vend_to_uuid.get(i) for t, i in d["vend_por_texto"].items()}

    # ── 2. Clientes ─────────────────────────────────────────────────────────
    filas_cli = [{
        "disapp_id": c["disapp_id"], "nombre": c["nombre"], "documento": c["documento"],
        "telefono": c["telefono"], "direccion": c["direccion"], "notas": c["notas"],
        "activo": c["activo"],
    } for c in d["clientes"].values()]
    print(f"\n[2] Clientes: {len(filas_cli)}  (documentos a NULL por duplicado: {d['dup_docs']})")
    disapp_to_uuid, doc_to_uuid = {}, {}
    if not dry:
        rep = upsert(db, "clientes", filas_cli, "disapp_id")
        for row in rep:
            disapp_to_uuid[row["disapp_id"]] = row["id"]
            if row.get("documento"):
                doc_to_uuid[row["documento"]] = row["id"]

    # ── 3. Creditos activos ─────────────────────────────────────────────────
    # Los créditos de supervisores YA NO se saltan: van al cobrador "CARTERA
    # <zona>" creado en el paso 1 (cartera VIP semanal de clientes reales).
    sup_ids = {i for i, n in d["vendedores"].items() if es_supervisor(n)}
    filas_cred, ref_to_uuid, ref_total_dias = [], {}, {}
    inc_cartera = skip_cli = 0
    for c in d["creditos"].values():
        if c["id_vendedor"] in sup_ids:
            inc_cartera += 1
        cli = disapp_to_uuid.get(c["id_cliente"]) if not dry else "?"
        if not dry and not cli:
            skip_cli += 1; continue
        td = c["cuotas"] or 1
        filas_cred.append({
            "disapp_credit_id": c["disapp_credit_id"], "disapp_credit_ref": c["ref"],
            "cliente_id": cli, "cobrador_id": id_vend_to_uuid.get(c["id_vendedor"]),
            "monto_prestado": c["monto_prestado"] or 1, "cuota_diaria": c["cuota"] or 1,
            "total_dias": td, "frecuencia": c["frecuencia"], "estado": "activo",
            "fecha_inicio": c["fecha"].isoformat() if c["fecha"] else None,
        })
        if c["ref"]:
            ref_total_dias[c["ref"]] = td
    print(f"\n[3] Creditos activos: {len(filas_cred)}  (cartera supervisor incluida: {inc_cartera}, sin cliente: {skip_cli})")
    if not dry:
        rep = upsert(db, "prestamos", filas_cred, "disapp_credit_id")
        for row in rep:
            if row.get("disapp_credit_ref"):
                ref_to_uuid[row["disapp_credit_ref"]] = row["id"]

    # ── 3b. Asignaciones cobrador<->cliente (RUTA del cobrador) ──────────────
    # SIN esto, el RLS (app_cobrador_tiene_cliente) deja la ruta del cobrador
    # VACÍA. Una asignación por cada (cobrador, cliente) de un crédito ACTIVO
    # (deduplicado). Un cliente con créditos activos de varios cobradores queda
    # en la ruta de todos ellos (la tabla lo permite: unique (cobrador,cliente)).
    pares = {(f["cobrador_id"], f["cliente_id"])
             for f in filas_cred if f["cobrador_id"] and f["cliente_id"]}
    filas_asig = [{"cobrador_id": cob, "cliente_id": cli, "activo": True} for cob, cli in pares]
    print(f"\n[3b] Asignaciones cobrador↔cliente: {len(filas_asig)}")
    if not dry:
        upsert(db, "asignaciones", filas_asig, "cobrador_id,cliente_id", rep=False)

    # ── 4. Stubs historicos (huerfanos) ─────────────────────────────────────
    synth_uuid = None
    filas_stub = []
    for ref, plist in d["huerfanos"].items():
        cuota_max = max((p["cuota_num"] or 1) for p in plist)
        td = max(cuota_max, max((p["total_cuotas"] or 1) for p in plist))
        doc = next((p["documento"] for p in plist if p["documento"]), None)
        vend = next((p["vendedor"] for p in plist if p["vendedor"]), None)
        fmin = min((p["fecha"] for p in plist if p["fecha"]), default=None)
        cuota = next((p["valor_cuota"] for p in plist if p["valor_cuota"]), None) or 1
        montoc = max((p["total_credito"] or 0) for p in plist) or (cuota * td)
        filas_stub.append({
            "_ref": ref, "_doc": doc, "_vend": vend,
            "disapp_credit_id": "STUB:" + ref, "disapp_credit_ref": ref,
            "monto_prestado": montoc, "cuota_diaria": cuota, "total_dias": td,
            "frecuencia": "diario", "estado": "finalizado",
            "fecha_inicio": fmin.isoformat() if fmin else None,
        })
        ref_total_dias[ref] = td
    sin_doc = sum(1 for s in filas_stub if not (s["_doc"] and (s["_doc"] in doc_to_uuid or dry)))
    print(f"\n[4] Stubs historicos: {len(filas_stub)}  (sin match de documento: {sin_doc} -> cliente sintetico)")
    if not dry and filas_stub:
        # cliente sintetico para huerfanos sin documento matcheable
        rep = upsert(db, "clientes", [{
            "disapp_id": "SYNTH:historico", "nombre": "(histórico sin identificar)",
            "documento": None, "activo": False,
        }], "disapp_id")
        synth_uuid = rep[0]["id"] if rep else None
        for s in filas_stub:
            s["cliente_id"] = doc_to_uuid.get(s["_doc"]) or synth_uuid
            s["cobrador_id"] = vtexto_to_uuid.get(s["_vend"])
        limpio = [{k: v for k, v in s.items() if not k.startswith("_")} for s in filas_stub]
        rep = upsert(db, "prestamos", limpio, "disapp_credit_id")
        for row in rep:
            if row.get("disapp_credit_ref"):
                ref_to_uuid[row["disapp_credit_ref"]] = row["id"]

    # ── 5. Pagos ────────────────────────────────────────────────────────────
    filas_pago, sin_prestamo, clamps = [], 0, 0
    for p in d["pagos"].values():
        pid = ref_to_uuid.get(p["ref"]) if not dry else "?"
        if not dry and not pid:
            sin_prestamo += 1; continue
        td = ref_total_dias.get(p["ref"], 10**6)
        dc = p["cuota_num"] or 1
        if dc > td:
            dc = td; clamps += 1
        if dc < 1:
            dc = 1
        filas_pago.append({
            "prestamo_id": pid, "dia_credito": dc, "monto": p["monto"] or 0.01,
            "registrado_por": vtexto_to_uuid.get(p["vendedor"]),
            "registrado_en": iso_ts(p["fecha"]) or iso_ts(dt.date.today()),
            "origen": "disapp_import", "importado_en": dt.datetime.now().isoformat(),
            "disapp_pago_id": p["id_pago"], "disapp_credit_ref": p["ref"],
        })
    print(f"\n[5] Pagos: {len(filas_pago)}  (sin prestamo: {sin_prestamo}, dia_credito clamped: {clamps})")
    if not dry:
        upsert(db, "pagos", filas_pago, "disapp_pago_id", ignore=True, rep=False)

    # ── 6. Reconstrucción de pagos NO-diarios ────────────────────────────────
    # DESPUÉS de los pagos reales, para conocer qué cuotas ya están cubiertas.
    # Siembra un ajuste (origen='ajuste_migracion') por la diferencia entre la
    # columna 'Pagos' de Disapp y lo importado de recaudos. Ver reconstruir_creditos.
    res = reconstruir_creditos(d)
    recon = [i for i in res.values() if i["sembrados"]]
    n_semb = sum(len(i["sembrados"]) for i in recon)
    monto_semb = round(sum(i["reconstruir"] for i in recon), 2)
    a_revision = [i for i in res.values() if i["revision"]]
    # Cartera pendiente global tras la reconstrucción = Σ falta (== Σ Saldo Pendiente
    # cuando la reconstrucción cierra; es el LIBRO COMPLETO y real, ~$94,2M — MÁS que
    # el tile "Capital en calle $68,5M" de Disapp, que esconde parte de su cartera).
    cartera_global = round(sum((i["falta_final"] or 0.0) for i in res.values()), 2)
    saldo_disapp = round(sum((i["saldo_pend"] or 0.0) for i in res.values()), 2)
    print(f"\n[6] Reconstrucción de pagos no-diarios:")
    print(f"      créditos activos evaluados:   {len(res):>8}")
    print(f"      créditos reconstruidos:       {len(recon):>8}")
    print(f"      pagos de ajuste sembrados:    {n_semb:>8}")
    print(f"      monto total sembrado:       $ {monto_semb:>14,.0f}")
    print(f"      créditos marcados a revisión: {len(a_revision):>8}  (Cuotas Pend. informativa)")
    print(f"      cartera pendiente global:   $ {cartera_global:>14,.0f}  (Σ Saldo Disapp: $ {saldo_disapp:,.0f})")

    # Reconciliación. Se separa el ÚNICO bucket que importa (dinero faltante real:
    # pagado < Pagos, imposible de sembrar sin borrar recaudos inmutables) de los
    # benignos: (a) pagado > Pagos = los recaudos traen pagos MÁS nuevos que el
    # snapshot 'Pagos' del Excel (mi dato es más actual → saldo un poco menor); y
    # (b) redondeo cuota*cuotas != Total. Ni (a) ni (b) son errores del pipeline.
    sub_cobrados = 0        # pagado < Pagos  → REVISAR (falta plata)
    mas_nuevos = 0          # pagado > Pagos  → OK (recaudos más recientes)
    exceso_nuevos = 0.0
    redondeo = 0            # cuota*cuotas != Total → OK (cosmético)
    filas_csv = []
    for info in res.values():
        fallos = _reconciliacion_identidades(info)
        motivos = info["revision"] + [f"{k}(esp={e:.0f},obt={o:.0f})" for k, e, o in fallos]
        pd_, pf_ = (info["pagos_disapp"] or 0.0), info["total_pagado_final"]
        if pf_ < pd_ - TOL_MONTO:
            sub_cobrados += 1
        elif pf_ > pd_ + TOL_MONTO:
            mas_nuevos += 1; exceso_nuevos += (pf_ - pd_)
        if any(k.startswith("cuota*cuotas") for k in info["revision"]):
            redondeo += 1
        if motivos:
            filas_csv.append((info["ref"], info["frecuencia"], " | ".join(motivos),
                              info["pagos_disapp"], info["total_pagado_final"],
                              info["saldo_pend"], info["falta_final"]))
    if filas_csv:
        p = os.path.join(out, "reconstruccion_revision.csv")
        with open(p, "w", encoding="utf-8") as f:
            f.write("credit_ref,frecuencia,motivo,pagos_disapp,pagado_final,saldo_disapp,falta_final\n")
            for ref, frec, mot, pd, pf, sd, ff in filas_csv:
                f.write(f'{ref},{frec},"{mot}",{pd or 0:.0f},{pf:.0f},{sd or 0:.0f},{ff or 0:.0f}\n')
        print(f"      🔴 sub-cobrados (pagado < Pagos, REVISAR): {sub_cobrados}")
        print(f"      ✔ recaudos más nuevos que snapshot (OK):  {mas_nuevos}  (+$ {exceso_nuevos:,.0f})")
        print(f"      ✔ redondeo cuota*cuotas != Total (OK):    {redondeo}")
        print(f"      revisión -> {p}  ({len(filas_csv)} filas; #3/CuotasPend informativas)")

    if not dry:
        filas_ajuste = _filas_ajuste(res, ref_to_uuid, dt.datetime.now().isoformat())
        print(f"      insertando {len(filas_ajuste)} pagos de ajuste...")
        upsert(db, "pagos", filas_ajuste, "disapp_pago_id", ignore=True, rep=False)

    print("\n" + ("=== DRY-RUN OK (no se escribio nada) ===" if dry
                  else "=== COMMIT COMPLETO ==="))

# ══ VALIDATE ════════════════════════════════════════════════════════════════
def count_rows(db, table, params=None):
    q = {"select": "id", "limit": 1}
    if params:
        q.update(params)
    url = db["url"] + "/rest/v1/" + table + "?" + urllib.parse.urlencode(q)
    req = urllib.request.Request(url, method="GET")
    req.add_header("apikey", db["key"]); req.add_header("Authorization", "Bearer " + db["key"])
    req.add_header("Prefer", "count=exact")
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            cr = resp.headers.get("Content-Range", "")
            return int(cr.split("/")[-1]) if "/" in cr else None
    except urllib.error.HTTPError:
        return None

def validate_sample(db, n):
    import random
    print("\n=== RECONCILIACION (contra la DB de prueba) ===")
    print(f"  clientes:            {count_rows(db, 'clientes')}")
    print(f"  cobradores/usuarios: {count_rows(db, 'usuarios')}")
    print(f"  prestamos totales:   {count_rows(db, 'prestamos')}")
    print(f"    · activos:         {count_rows(db, 'prestamos', {'estado':'eq.activo'})}")
    print(f"    · finalizados:     {count_rows(db, 'prestamos', {'estado':'eq.finalizado'})}")
    print(f"  pagos:               {count_rows(db, 'pagos')}")

    print(f"\n=== VALIDACION: {n} creditos ACTIVOS al azar (compará contra Disapp) ===")
    # Muestreo sobre creditos activos (los que definen el carton de go-live).
    act = get_rows(db, "prestamos", "id,cliente_id,cuota_diaria,total_dias,disapp_credit_ref",
                   {"estado": "eq.activo"})
    for pr in random.sample(act, min(n, len(act))):
        cli = get_rows(db, "clientes", "nombre", {"id": f"eq.{pr['cliente_id']}", "limit": 1})
        nombre = cli[0]["nombre"] if cli else "?"
        pagos = get_rows(db, "pagos", "monto", {"prestamo_id": f"eq.{pr['id']}", "anulado": "eq.false"})
        pagado = sum(float(x["monto"]) for x in pagos)
        total = float(pr["cuota_diaria"]) * int(pr["total_dias"])
        print(f"  {nombre[:30]:30} {str(pr['disapp_credit_ref'] or ''):>12}  "
              f"pagado ${pagado:>11,.0f} / total ${total:>11,.0f}  "
              f"saldo ${max(0,total-pagado):>11,.0f}  ({len(pagos)} pagos)")

# ══ PROBE (cuadre dirigido, antes del masivo) ═══════════════════════════════
def probe(src, refs):
    """Reconstruye SOLO los créditos indicados y muestra, lado a lado, lo que dice
    Disapp (Saldo Pendiente / Cuotas Pend.) vs lo que da el cartón tras sembrar.
    No toca la DB. Uso: --probe PRD0003208827,PRD0002535750."""
    only = {r.strip() for r in refs.split(",") if r.strip()}
    d = consolidar(src)
    res = reconstruir_creditos(d, only_refs=only)
    print("\n=== PROBE (cuadre dirigido, sin escribir) ===")
    faltantes = only - set(res.keys())
    for r in sorted(faltantes):
        print(f"  ⚠ {r}: no es un crédito ACTIVO del Excel (no reconstruible)")
    for ref in sorted(res.keys()):
        i = res[ref]
        cuota = i["cuota"] or 0
        pagadas = int(i["total_pagado_final"] // cuota) if cuota else 0
        toca = "SÍ" if i["sembrados"] else "no (ya completo)"
        print(f"\n  {ref}  [{i['frecuencia']}]  cuota ${cuota:,.0f} x {i['total_dias']}")
        print(f"    recaudos reales importados : $ {i['ya']:>12,.0f}")
        print(f"    Pagos (Disapp)             : $ {(i['pagos_disapp'] or 0):>12,.0f}")
        print(f"    ¿reconstruye?              : {toca}  (+$ {i['reconstruir']:,.0f}, {len(i['sembrados'])} pagos)")
        print(f"    pagado final (cartón)      : $ {i['total_pagado_final']:>12,.0f}")
        print(f"    ── SALDO ──  cartón: $ {(i['falta_final'] or 0):>12,.0f}   Disapp: $ {(i['saldo_pend'] or 0):>12,.0f}")
        print(f"    ── CUOTAS ─  cartón pagadas: {pagadas:>4}   Disapp pagadas: {(i['total_dias'] - i['cuotas_pend']) if i['cuotas_pend'] is not None else '?'}")
        fallos = _reconciliacion_identidades(i)
        if fallos:
            print(f"    identidades que fallan: " + ", ".join(k for k, *_ in fallos))
        else:
            print(f"    ✓ identidades duras (#1 suma==Pagos, #2 saldo==Saldo) OK")

# ══ main ════════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser(description="Empalme Disapp -> Presta Ya")
    ap.add_argument("--src", default=DEFAULT_SRC)
    ap.add_argument("--out", default=None)
    ap.add_argument("--env-file", default=".env.prueba")
    ap.add_argument("--audit", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--commit", action="store_true")
    ap.add_argument("--prod", action="store_true")
    ap.add_argument("--validate-sample", type=int, default=0)
    ap.add_argument("--probe", default=None,
                    help="Cuadre dirigido de refs (coma-sep), sin escribir. Ej: PRD0003208827,PRD0002535750")
    a = ap.parse_args()
    out = a.out or os.path.join(a.src, "_reportes")
    if not os.path.isdir(a.src):
        sys.exit(f"No existe la carpeta de datos: {a.src}")

    if a.probe:
        probe(a.src, a.probe)
        return

    if a.audit:
        audit(a.src, out)
        return

    # write path (dry-run por defecto)
    dry = not a.commit
    db = None
    if a.commit or a.validate_sample:
        env = load_env(a.env_file)
        url = env.get("SUPABASE_URL") or env.get("NEXT_PUBLIC_SUPABASE_URL")
        key = env.get("SUPABASE_SERVICE_ROLE_KEY")
        if not url or not key:
            sys.exit(f"Faltan SUPABASE_URL/SERVICE_ROLE_KEY en {a.env_file}")
        host = urllib.parse.urlparse(url).netloc
        es_prueba = any(x in os.path.basename(a.env_file).lower() for x in ("prueba", "test"))
        print(f"Destino: {host}  (.env: {a.env_file})")
        if a.commit and not es_prueba and not a.prod:
            sys.exit("ABORTADO: el .env no parece de prueba y no pasaste --prod.")
        db = {"url": url.rstrip("/"), "key": key, "host": host}

    if a.validate_sample and not a.commit:
        validate_sample(db, a.validate_sample)
        return

    commit_import(a.src, out, db, dry)
    if a.commit and a.validate_sample:
        validate_sample(db, a.validate_sample)

if __name__ == "__main__":
    main()
