# -*- coding: utf-8 -*-
"""
Cobros cargados en la app que Disapp NO tiene y que valen EXACTAMENTE el saldo.

Regla del negocio (Carlos, 06-08): **Disapp es la verdad** — el personal trabaja
con esa aplicación y la información tiene que coincidir.

Por qué "exactamente el saldo" es la señal: la app tiene un tope anti-sobre-pago
que RECORTA al saldo lo que se tipea de más. Cualquier monto que se pase produce
esa cifra exacta. Nadie elige ese número a mano — pero un cliente que termina de
pagar SÍ paga justo el saldo, así que la coincidencia sola no alcanza: por eso se
exige ADEMÁS que Disapp no tenga ese cobro. Si el cliente hubiera pagado de
verdad, el cobrador lo habría cargado en Disapp como cargó los otros 1.248.

NO se anulan los cobros de la app que Disapp tampoco tiene pero que NO son el
saldo: esos son cobros normales que el cobrador registró acá y no allá — la plata
existe y el cliente la entregó.

  Ver:      python scripts/_anular-pagos-clampeados-0805.py
  Anular:   python scripts/_anular-pagos-clampeados-0805.py --commit
"""
import json
import os
import re
import ssl
import sys
import datetime as dt
from urllib.parse import unquote

import openpyxl
import pg8000.dbapi

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORT = r"C:\Users\Carlos\migracion\creditos_2026-08-06_04-07.xlsx"
COMMIT = "--commit" in sys.argv
# Solo los del piloto en adelante: lo viejo es otra historia y se revisa aparte.
DESDE = dt.date(2026, 8, 5)
TOL = 1.0
MOTIVO = ("Monto tipeado por encima del saldo: el servidor lo recortó al saldo entero del "
          "crédito. Disapp no registra ese cobro, así que no corresponde a efectivo "
          "recibido. Se anula para que la caja del cobrador refleje lo que cobró de verdad.")


def num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def conectar():
    with open(os.path.join(RAIZ, ".env.local"), encoding="utf-8") as fh:
        url = next(l.split("=", 1)[1].strip().strip('"').strip("'")
                   for l in fh if l.startswith("SUPABASE_DB_URL="))
    m = re.match(r"postgres(?:ql)?://([^:]+):([^@]+)@([^:/]+):(\d+)/(.+)", url)
    usr, pw, host, port, base = m.groups()
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return pg8000.dbapi.connect(user=unquote(usr), password=unquote(pw), host=host,
                                port=int(port), database=base.split("?")[0], ssl_context=ctx)


def money(n) -> str:
    return "$ " + f"{round(float(n or 0)):,}".replace(",", ".")


def main() -> None:
    wb = openpyxl.load_workbook(EXPORT, read_only=True, data_only=True)
    ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    iRef, iPag, iSal = h.index("Crédito #"), h.index("Pagos"), h.index("Saldo Pendiente")
    disapp = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        ref = str(r[iRef] or "").strip()
        if ref:
            disapp[ref] = {"pagos": num(r[iPag]), "saldo": num(r[iSal])}
    wb.close()

    cn = conectar()
    cur = cn.cursor()
    cur.execute("""
        SELECT pg.id, cl.nombre, pr.disapp_credit_ref, pg.monto, pr.cuota_diaria,
               pr.pagado_acum, (pg.registrado_en AT TIME ZONE 'America/Montevideo') t,
               u.nombre
          FROM pagos pg
          JOIN prestamos pr ON pr.id = pg.prestamo_id
          JOIN clientes cl ON cl.id = pr.cliente_id
          LEFT JOIN usuarios u ON u.id = pg.registrado_por
         WHERE pg.origen IS NULL AND pg.anulado = false
           AND pr.disapp_credit_ref IS NOT NULL
           AND (pg.registrado_en AT TIME ZONE 'America/Montevideo')::date >= %s
         ORDER BY pg.monto DESC
    """, (DESDE,))

    anular, normales = [], []
    for pid, nom, ref, monto, cuota, acum, t, cob in cur.fetchall():
        d = disapp.get(ref)
        if not d:
            continue                                   # Disapp ya lo cerró: otro caso
        monto, acum = float(monto), float(acum)
        if acum - d["pagos"] < TOL:
            continue                                   # coincidimos con Disapp: nada que ver
        fila = {"pago_id": str(pid), "cliente": nom, "ref": ref, "monto": monto,
                "cuota": float(cuota), "saldoDisapp": d["saldo"], "cuando": str(t),
                "cobrador": cob}
        # La firma: el cobro vale EXACTAMENTE el saldo que Disapp dice pendiente.
        (anular if abs(monto - d["saldo"]) < TOL else normales).append(fila)

    print(f"{'='*104}\n  COBROS EN LA APP QUE DISAPP NO TIENE   "
          f"{'🔴 ANULANDO' if COMMIT else '🟡 SOLO MIRANDO'}   (desde {DESDE})\n{'='*104}")
    print(f"\n  ✗ VALEN EXACTAMENTE EL SALDO (recorte del servidor → tipeo de más): {len(anular)}")
    for a in anular:
        print(f"     {a['cliente'][:28]:28s} {a['ref']:16s} {money(a['monto']):>11s}  "
              f"cuota {money(a['cuota']):>9s}  {a['cuando'][:16]}  {str(a['cobrador'])[:18]}")
    print(f"     total: {money(sum(a['monto'] for a in anular))}")
    print(f"\n  ✓ NO son el saldo — cobros normales que Disapp no tiene. SE DEJAN: {len(normales)}")
    for x in normales:
        print(f"     {x['cliente'][:28]:28s} {x['ref']:16s} {money(x['monto']):>11s}  "
              f"(saldo Disapp {money(x['saldoDisapp'])})  {str(x['cobrador'])[:18]}")

    if not COMMIT:
        print("\n🟡 No se tocó nada. Confirmá con los cobradores y corré con --commit.\n")
        return
    if not anular:
        print("\nNada que anular.\n")
        return

    cur.execute("SELECT id FROM usuarios WHERE rol='admin' AND nombre ILIKE 'Mauricio%' LIMIT 1")
    fila = cur.fetchone()
    if not fila:
        cur.execute("SELECT id FROM usuarios WHERE rol='admin' ORDER BY nombre LIMIT 1")
        fila = cur.fetchone()
    if not fila:
        print("🔴 No hay admin para firmar la anulación. Abortado.")
        return

    log = os.path.join(RAIZ, "scripts", f"_revert_clampeados_{dt.date.today():%Y%m%d}.json")
    with open(log, "w", encoding="utf-8") as fh:
        json.dump(anular, fh, ensure_ascii=False, indent=1, default=str)

    cur.execute("""
        UPDATE pagos
           SET anulado = true, anulado_en = now(), anulado_por = %s, motivo_anulacion = %s
         WHERE id = ANY(%s) AND anulado = false
    """, (fila[0], MOTIVO, [a["pago_id"] for a in anular]))
    cn.commit()
    print(f"\n  ✓ anulados: {cur.rowcount}    log de reversa: {log}")

    for quien in sorted({str(a["cobrador"]) for a in anular}):
        cur.execute("""
            SELECT COALESCE(sum(pg.monto), 0), count(*)
              FROM pagos pg JOIN usuarios u ON u.id = pg.registrado_por
             WHERE u.nombre = %s AND pg.origen IS NULL AND pg.anulado = false
               AND (pg.registrado_en AT TIME ZONE 'America/Montevideo')::date = '2026-08-05'
        """, (quien,))
        t, n = cur.fetchone()
        print(f"  caja del 05-08 de {quien}: {money(t)} en {n} cobros")

    cur.close()
    cn.close()


if __name__ == "__main__":
    main()
