# -*- coding: utf-8 -*-
"""
Alinea la FORMA del crédito (cuota, cantidad de cuotas, frecuencia) con Disapp.

Regla de Carlos: Disapp manda — el personal trabaja con esa aplicación y frente al
cliente gana lo que diga Disapp. Si acá la cuota es $600 y allá $500, el cobrador
le pide de más y el cliente discute con razón.

No toca un solo peso del libro de pagos: solo corrige las CONDICIONES del crédito.
El estado del cartón se deriva de los pagos, así que al cambiar la cuota el avance
del cliente se recalcula solo y queda coherente con lo que él ve en su cartón.

  Ver:      python scripts/_alinear-forma-credito-0806.py
  Aplicar:  python scripts/_alinear-forma-credito-0806.py --commit
"""
import json
import os
import re
import ssl
import sys
import datetime as dt
from urllib.parse import unquote

import openpyxl
import pg8000.dbapi

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORT = r"C:\Users\Carlos\migracion\creditos_2026-08-06_04-07.xlsx"
COMMIT = "--commit" in sys.argv
TOL = 1.0

MODALIDAD = {
    "diaria": "diario",
    "semanal": "semanal",
    "quincenal": "quincenal",
    "mensual": "mensual",
}


def num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def conectar():
    with open(os.path.join(RAIZ, ".env.local"), encoding="utf-8") as fh:
        url = next(l.split("=", 1)[1].strip().strip('"').strip("'")
                   for l in fh if l.startswith("SUPABASE_DB_URL="))
    m = re.match(r"postgres(?:ql)?://([^:]+):([^@]+)@([^:/]+):(\d+)/(.+)", url)
    usr, pw, host, port, base = m.groups()
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return pg8000.dbapi.connect(user=unquote(usr), password=unquote(pw), host=host,
                                port=int(port), database=base.split("?")[0], ssl_context=ctx)


def money(n) -> str:
    return "$ " + f"{round(float(n or 0)):,}".replace(",", ".")


def main() -> None:
    wb = openpyxl.load_workbook(EXPORT, read_only=True, data_only=True)
    ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {k: h.index(k) for k in ["Crédito #", "Valor Cuota", "Cuotas", "Modalidad", "Total c/ Intereses"]}
    disapp = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        ref = str(r[idx["Crédito #"]] or "").strip()
        if not ref:
            continue
        disapp[ref] = {
            "cuota": num(r[idx["Valor Cuota"]]),
            "dias": int(num(r[idx["Cuotas"]])),
            "frec": MODALIDAD.get(str(r[idx["Modalidad"]] or "").strip().lower()),
            "total": num(r[idx["Total c/ Intereses"]]),
        }
    wb.close()

    cn = conectar()
    cur = cn.cursor()
    cur.execute("""
        SELECT pr.id, pr.disapp_credit_ref, pr.cuota_diaria, pr.total_dias,
               COALESCE(pr.frecuencia,'diario'), pr.pagado_acum, cl.nombre, u.nombre
          FROM prestamos pr JOIN clientes cl ON cl.id = pr.cliente_id
          LEFT JOIN usuarios u ON u.id = pr.cobrador_id
         WHERE pr.estado = 'activo' AND pr.disapp_credit_ref IS NOT NULL
    """)
    difs = []
    for pid, ref, cuota, dias, frec, pagado, cliente, cobrador in cur.fetchall():
        d = disapp.get(ref)
        if not d:
            continue
        cambios = {}
        # ⚠️ Solo diferencias que cambian lo que se le PIDE al cliente. El export
        # trae la cuota redondeada a 2 decimales y la nuestra es la exacta
        # (heredada: 8.425/24 = 351,0416…), así que comparar al centavo marca 46
        # créditos que en la calle son idénticos. Un peso es el mínimo cobrable.
        if d["cuota"] > 0 and abs(float(cuota) - d["cuota"]) >= 1:
            cambios["cuota_diaria"] = d["cuota"]
        if d["dias"] > 0 and int(dias) != d["dias"]:
            cambios["total_dias"] = d["dias"]
        if d["frec"] and frec != d["frec"]:
            cambios["frecuencia"] = d["frec"]
        if not cambios:
            continue
        nuevoTotal = (cambios.get("cuota_diaria", float(cuota))) * (cambios.get("total_dias", int(dias)))
        difs.append({
            "id": str(pid), "ref": ref, "cliente": cliente, "cobrador": cobrador,
            "antes": {"cuota": float(cuota), "dias": int(dias), "frecuencia": frec,
                      "total": float(cuota) * int(dias)},
            "despues": {**{"cuota": d["cuota"], "dias": d["dias"], "frecuencia": d["frec"] or frec},
                        "total": nuevoTotal},
            "cambios": cambios, "pagado": float(pagado), "totalDisapp": d["total"],
        })

    print(f"{'='*104}\n  FORMA DEL CRÉDITO vs DISAPP   {'🔴 APLICANDO' if COMMIT else '🟡 SOLO MIRANDO'}\n{'='*104}")
    print(f"\n  créditos con la forma distinta: {len(difs)}\n")
    for d in difs:
        a, p = d["antes"], d["despues"]
        print(f"  {d['cliente'][:30]:30s} {d['ref']:16s}  ({str(d['cobrador'])[:20]})")
        print(f"     nosotros: {money(a['cuota']):>10s} × {a['dias']:>3d} {a['frecuencia']:10s} = {money(a['total'])}")
        print(f"     Disapp:   {money(p['cuota']):>10s} × {p['dias']:>3d} {p['frecuencia']:10s} = {money(p['total'])}"
              f"   (su 'Total c/Int' dice {money(d['totalDisapp'])})")
        print(f"     pagado hasta hoy: {money(d['pagado'])}")
        if p["cuota"] < a["cuota"]:
            print(f"     ⚠️ le estábamos pidiendo {money(a['cuota'] - p['cuota'])} de MÁS por cuota")
        print()

    if not COMMIT:
        print("🟡 No se tocó nada. Aplicar con --commit.\n")
        return
    if not difs:
        print("Nada que alinear.\n")
        return

    log = os.path.join(RAIZ, "scripts", f"_revert_forma_{dt.date.today():%Y%m%d}.json")
    with open(log, "w", encoding="utf-8") as fh:
        json.dump(difs, fh, ensure_ascii=False, indent=1, default=str)

    n = 0
    for d in difs:
        sets, vals = [], []
        for col, v in d["cambios"].items():
            sets.append(f"{col} = %s")
            vals.append(v)
        vals.append(d["id"])
        cur.execute(f"UPDATE prestamos SET {', '.join(sets)} WHERE id = %s AND estado='activo'", vals)
        n += cur.rowcount
    cn.commit()
    print(f"  ✓ créditos alineados: {n}    log de reversa: {log}\n")
    cur.close()
    cn.close()


if __name__ == "__main__":
    main()
