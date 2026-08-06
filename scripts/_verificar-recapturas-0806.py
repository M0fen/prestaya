# -*- coding: utf-8 -*-
"""
¿Las 301 "re-capturas" son duplicados de verdad, o plata que solo vive en Disapp?

El empalme descarta las filas del export que caen sobre créditos que la app ya
maneja y son posteriores al arranque del piloto: asume que el cobrador las cargó
en los DOS lados y que la app es la verdad. Si esa suposición falla —el cobrador
cobró y lo anotó SOLO en Disapp— el pago no entra nunca a Presta Ya, y mañana la
app le va a pedir esa cuota otra vez al cliente. Doble cobro.

Este script mira, fila por fila, si existe un pago NATIVO en la app para el mismo
crédito el mismo día. Solo LECTURA.
"""
import os
import re
import ssl
import datetime as dt
from collections import defaultdict
from urllib.parse import unquote

import openpyxl
import pg8000.dbapi

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORT = r"C:\Users\Carlos\migracion\recaudos_2026-08-06_02-15.xlsx"
PILOTO_DESDE = dt.date(2026, 8, 5)


def conectar():
    with open(os.path.join(RAIZ, ".env.local"), encoding="utf-8") as fh:
        url = next(l.split("=", 1)[1].strip().strip('"').strip("'")
                   for l in fh if l.startswith("SUPABASE_DB_URL="))
    m = re.match(r"postgres(?:ql)?://([^:]+):([^@]+)@([^:/]+):(\d+)/(.+)", url)
    usr, pw, host, port, base = m.groups()
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return pg8000.dbapi.connect(user=unquote(usr), password=unquote(pw), host=host,
                                port=int(port), database=base.split("?")[0], ssl_context=ctx)


def money(n) -> str:
    return "$ " + f"{round(float(n or 0)):,}".replace(",", ".")


def fecha_de(v) -> dt.date | None:
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s[:10], fmt).date()
        except ValueError:
            pass
    return None


def main() -> None:
    wb = openpyxl.load_workbook(EXPORT, read_only=True, data_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    iId, iRef = hdr.index("ID Pago"), hdr.index("Ref. Crédito")
    iRec, iFec, iVen = hdr.index("Recaudo"), hdr.index("Fecha Pago"), hdr.index("Vendedor")
    iCli = hdr.index("Cliente")
    filas = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        ref = str(r[iRef] or "").strip()
        f = fecha_de(r[iFec])
        if not ref or not f:
            continue
        filas.append({"id": str(r[iId]), "ref": ref, "fecha": f,
                      "monto": float(r[iRec] or 0), "vend": str(r[iVen] or ""),
                      "cliente": str(r[iCli] or "")})
    wb.close()
    print(f"Export: {len(filas)} recaudos  ({min(f['fecha'] for f in filas)} → {max(f['fecha'] for f in filas)})")

    cn = conectar()
    cur = cn.cursor()

    # Créditos de la app indexados por su ref de Disapp.
    cur.execute("SELECT id, disapp_credit_ref, cobrador_id FROM prestamos WHERE disapp_credit_ref IS NOT NULL")
    por_ref = {r[1]: {"id": r[0], "cob": r[2]} for r in cur.fetchall()}

    # Cobradores "vivos": los que ya registran en la app desde el piloto.
    cur.execute("""
        SELECT DISTINCT registrado_por FROM pagos
         WHERE origen IS NULL AND anulado = false AND registrado_en >= '2026-08-05T03:00:00Z'
    """)
    vivos = {r[0] for r in cur.fetchall() if r[0]}

    # Pagos NATIVOS por (prestamo, día UY).
    cur.execute("""
        SELECT prestamo_id, (registrado_en AT TIME ZONE 'America/Montevideo')::date, count(*), sum(monto)
          FROM pagos
         WHERE origen IS NULL AND anulado = false AND registrado_en >= '2026-08-05T03:00:00Z'
         GROUP BY 1,2
    """)
    nativo = {(r[0], r[1]): {"n": r[2], "monto": float(r[3] or 0)} for r in cur.fetchall()}

    # Pagos YA importados (para no contar los que ya entraron).
    cur.execute("SELECT disapp_pago_id FROM pagos WHERE disapp_pago_id IS NOT NULL")
    ya = {str(r[0]) for r in cur.fetchall()}

    dup, solo_disapp, sin_credito, ya_estan = [], [], [], 0
    for f in filas:
        if f["id"] in ya:
            ya_estan += 1
            continue
        p = por_ref.get(f["ref"])
        if not p:
            sin_credito.append(f)
            continue
        # Solo aplica la regla de re-captura si el crédito es de un cobrador vivo.
        if f["fecha"] < PILOTO_DESDE or p["cob"] not in vivos:
            continue
        if (p["id"], f["fecha"]) in nativo:
            dup.append(f)
        else:
            solo_disapp.append(f)

    print(f"\nya importados (dedupe por ID Pago):        {ya_estan}")
    print(f"sin crédito en la app (ref desconocida):   {len(sin_credito)}  {money(sum(x['monto'] for x in sin_credito))}")
    print("\n═══ LAS 'RE-CAPTURAS' QUE EL EMPALME DESCARTA ═══")
    print(f"  ✅ duplicados REALES (hay pago nativo ese día):  {len(dup):5d}  {money(sum(x['monto'] for x in dup))}")
    print(f"  🔴 SOLO EN DISAPP (la app NO tiene ese cobro):   {len(solo_disapp):5d}  {money(sum(x['monto'] for x in solo_disapp))}")

    if solo_disapp:
        print("\n  ⚠️ Si no entran, mañana la app les vuelve a pedir esa cuota a estos clientes:")
        por_vend = defaultdict(lambda: [0, 0.0])
        for x in solo_disapp:
            por_vend[x["vend"]][0] += 1
            por_vend[x["vend"]][1] += x["monto"]
        for v, (n, m) in sorted(por_vend.items(), key=lambda kv: -kv[1][1]):
            print(f"     {v[:38]:38s} {n:4d} cobros  {money(m)}")
        print("\n  Ejemplos:")
        for x in sorted(solo_disapp, key=lambda y: -y["monto"])[:8]:
            print(f"     {x['cliente'][:32]:32s} {x['ref']:16s} {x['fecha']}  {money(x['monto'])}")

    cur.close()
    cn.close()


if __name__ == "__main__":
    main()
