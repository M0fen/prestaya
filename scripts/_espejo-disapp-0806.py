# -*- coding: utf-8 -*-
"""
¿Somos espejo de Disapp? Respuesta honesta, crédito por crédito y peso por peso.

Compara contra el EXPORT (no contra el tablero: el tablero aplica filtros propios
y ni siquiera coincide con su propio export). Tres preguntas:
  1. ¿Están TODOS los créditos activos de Disapp? ¿Sobra alguno?
  2. ¿Cuánto pagó cada uno según ellos vs según nosotros?
  3. ¿Están todos los clientes?

Solo LECTURA.
"""
import os
import re
import ssl
from urllib.parse import unquote

import openpyxl
import pg8000.dbapi

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORT_CRED = r"C:\Users\Carlos\migracion\creditos_2026-08-06_04-07.xlsx"
EXPORT_CLI = r"C:\Users\Carlos\migracion\clientes_2026-07-20_12-51.xlsx"
TOL = 1.0


def num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def conectar():
    with open(os.path.join(RAIZ, ".env.local"), encoding="utf-8") as fh:
        url = next(l.split("=", 1)[1].strip().strip('"').strip("'")
                   for l in fh if l.startswith("SUPABASE_DB_URL="))
    m = re.match(r"postgres(?:ql)?://([^:]+):([^@]+)@([^:/]+):(\d+)/(.+)", url)
    usr, pw, host, port, base = m.groups()
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return pg8000.dbapi.connect(user=unquote(usr), password=unquote(pw), host=host,
                                port=int(port), database=base.split("?")[0], ssl_context=ctx)


def money(n) -> str:
    return "$ " + f"{round(float(n or 0)):,}".replace(",", ".")


def pct(a, b) -> str:
    return f"{100.0*a/b:.2f}%" if b else "—"


def main() -> None:
    wb = openpyxl.load_workbook(EXPORT_CRED, read_only=True, data_only=True)
    ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    iRef, iPag, iCli, iVen = h.index("Crédito #"), h.index("Pagos"), h.index("Cliente"), h.index("Vendedor")
    exp = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        ref = str(r[iRef] or "").strip()
        if ref:
            exp[ref] = {"pagos": num(r[iPag]), "cliente": str(r[iCli] or ""), "vend": str(r[iVen] or "")}
    wb.close()

    wb = openpyxl.load_workbook(EXPORT_CLI, read_only=True, data_only=True)
    ws = wb.active
    h2 = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    iId = h2.index("ID")
    cli_exp = {str(r[iId]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[iId]}
    wb.close()

    cn = conectar()
    cur = cn.cursor()
    cur.execute("""
        SELECT disapp_credit_ref, pagado_acum, disapp_credit_id
          FROM prestamos WHERE estado = 'activo'
    """)
    filas = cur.fetchall()
    nuestros = {r[0]: float(r[1] or 0) for r in filas if r[0]}
    nativos = sum(1 for r in filas if not r[0])

    print(f"\n{'='*98}\n  ESPEJO CONTRA EL EXPORT DE DISAPP (creditos_2026-08-06_04-07)\n{'='*98}")

    # ── 1. Créditos ────────────────────────────────────────────────────────
    faltan = [r for r in exp if r not in nuestros]
    sobran = [r for r in nuestros if r not in exp]
    print("\n── 1. ¿ESTÁN TODOS LOS CRÉDITOS ACTIVOS? ──")
    print(f"  activos en el export de Disapp:        {len(exp)}")
    print(f"  de esos, ACTIVOS en Presta Ya:         {len(exp) - len(faltan)}   ({pct(len(exp)-len(faltan), len(exp))})")
    print(f"  ✗ del export que NOS FALTAN:           {len(faltan)}")
    print(f"  + activos nuestros que Disapp NO lista: {len(sobran)}  (no se cierran a propósito)")
    print(f"  + nacidos en Presta Ya (sin ref):       {nativos}")
    for r in faltan[:5]:
        print(f"      falta: {r}  {exp[r]['cliente'][:30]}")

    # ── 2. Plata por crédito ───────────────────────────────────────────────
    exactos, cortos, pasados = [], [], []
    for ref, e in exp.items():
        if ref not in nuestros:
            continue
        gap = round(e["pagos"] - nuestros[ref], 2)
        (exactos if abs(gap) <= TOL else (cortos if gap > 0 else pasados)).append((ref, gap, e))
    total = len(exactos) + len(cortos) + len(pasados)
    print("\n── 2. ¿CUÁNTO PAGÓ CADA UNO? (columna 'Pagos' de Disapp vs nuestro acumulado) ──")
    print(f"  IDÉNTICOS al peso:      {len(exactos):5d}  ({pct(len(exactos), total)})")
    print(f"  nos falta registrar:    {len(cortos):5d}   {money(sum(g for _, g, _ in cortos))}")
    print(f"  tenemos de más:         {len(pasados):5d}   {money(abs(sum(g for _, g, _ in pasados)))}")
    if cortos:
        print("\n  Donde NOS FALTA plata (el cliente pagó y no lo tenemos → se le puede pedir de nuevo):")
        for ref, g, e in sorted(cortos, key=lambda x: -x[1])[:8]:
            print(f"     {e['cliente'][:30]:30s} {ref:16s} falta {money(g):>12s}  ({e['vend'][:22]})")
    if pasados:
        print("\n  Donde tenemos DE MÁS (le acreditamos algo que Disapp no tiene):")
        for ref, g, e in sorted(pasados, key=lambda x: x[1])[:8]:
            print(f"     {e['cliente'][:30]:30s} {ref:16s} sobra {money(abs(g)):>12s}  ({e['vend'][:22]})")

    # ── 3. Clientes ────────────────────────────────────────────────────────
    cur.execute("SELECT count(*) FROM clientes")
    cli_total = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM clientes WHERE disapp_id IS NOT NULL")
    cli_con_ref = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM clientes WHERE disapp_id IS NULL")
    cli_nativos = cur.fetchone()[0]
    print("\n── 3. ¿ESTÁN TODOS LOS CLIENTES? ──")
    print(f"  clientes en el export:        {len(cli_exp)}")
    print(f"  clientes en Presta Ya:        {cli_total}   (de Disapp {cli_con_ref} · nacidos acá {cli_nativos})")

    # ── 4. Totales de plata ────────────────────────────────────────────────
    sum_exp = sum(e["pagos"] for r, e in exp.items() if r in nuestros)
    sum_nos = sum(nuestros[r] for r in exp if r in nuestros)
    print("\n── 4. LA PLATA, EN TOTAL (sobre los créditos que ambos tenemos activos) ──")
    print(f"  recaudado según Disapp:   {money(sum_exp)}")
    print(f"  recaudado según nosotros: {money(sum_nos)}")
    d = sum_nos - sum_exp
    print(f"  diferencia:               {money(d)}   ({pct(abs(d), sum_exp)} del total)")

    print(f"\n{'='*98}")
    veredicto = (
        "ESPEJO EXACTO" if not faltan and not cortos and not pasados
        else f"ESPEJO AL {pct(len(exactos), total)} — quedan {len(cortos)+len(pasados)} créditos con diferencia"
    )
    print(f"  VEREDICTO: {veredicto}")
    print(f"{'='*98}\n")
    cur.close()
    cn.close()


if __name__ == "__main__":
    main()
