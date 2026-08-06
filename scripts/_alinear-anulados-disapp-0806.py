# -*- coding: utf-8 -*-
"""
Anula acá los cobros que DISAPP anuló allá.

Regla del negocio (Carlos, 06-08): Disapp es la verdad — el personal trabaja con
esa aplicación y la información tiene que coincidir.

Qué pasa: un cobro entra al export, lo importamos, y más tarde Disapp lo REVIERTE
(cheque sin fondo, corrección, error de carga). Su columna 'Pagos' baja, la
nuestra no, y ese crédito queda para siempre mostrando más pagado de lo real —
o sea, le cobramos de menos al cliente.

Cómo se detecta sin adivinar: la diferencia contra Disapp tiene que coincidir
EXACTAMENTE con UN pago nuestro que tenga `disapp_pago_id` (vino de allá). Si el
excedente es un residuo de varios pagos, NO se toca: no hay forma de saber cuál.

Los pagos no se borran: se anulan con quién y por qué (libro inmutable).

  Ver:      python scripts/_alinear-anulados-disapp-0806.py
  Aplicar:  python scripts/_alinear-anulados-disapp-0806.py --commit
"""
import json
import os
import re
import ssl
import sys
import datetime as dt
from urllib.parse import unquote

import openpyxl
import pg8000.dbapi

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORT = r"C:\Users\Carlos\migracion\creditos_2026-08-06_04-07.xlsx"
COMMIT = "--commit" in sys.argv
TOL = 1.0
MOTIVO = ("Cobro revertido en Disapp: su columna 'Pagos' ya no lo incluye. Se alinea "
          "el libro para que el saldo del cliente sea el mismo en los dos sistemas.")


def num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def conectar():
    with open(os.path.join(RAIZ, ".env.local"), encoding="utf-8") as fh:
        url = next(l.split("=", 1)[1].strip().strip('"').strip("'")
                   for l in fh if l.startswith("SUPABASE_DB_URL="))
    m = re.match(r"postgres(?:ql)?://([^:]+):([^@]+)@([^:/]+):(\d+)/(.+)", url)
    usr, pw, host, port, base = m.groups()
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return pg8000.dbapi.connect(user=unquote(usr), password=unquote(pw), host=host,
                                port=int(port), database=base.split("?")[0], ssl_context=ctx)


def money(n) -> str:
    return "$ " + f"{round(float(n or 0)):,}".replace(",", ".")


def main() -> None:
    wb = openpyxl.load_workbook(EXPORT, read_only=True, data_only=True)
    ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    iRef, iPag = h.index("Crédito #"), h.index("Pagos")
    disapp = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        ref = str(r[iRef] or "").strip()
        if ref:
            disapp[ref] = num(r[iPag])
    wb.close()

    cn = conectar()
    cur = cn.cursor()
    cur.execute("""
        SELECT pr.disapp_credit_ref, pr.pagado_acum, cl.nombre
          FROM prestamos pr JOIN clientes cl ON cl.id = pr.cliente_id
         WHERE pr.estado = 'activo' AND pr.disapp_credit_ref IS NOT NULL
    """)
    anular, sin_match = [], []
    for ref, acum, nom in cur.fetchall():
        if ref not in disapp:
            continue
        dif = round(float(acum) - disapp[ref], 2)
        if dif <= TOL:
            continue
        # ¿La diferencia es EXACTAMENTE un pago importado que sigue vivo?
        cur.execute("""
            SELECT pg.id, pg.monto, pg.disapp_pago_id,
                   (pg.registrado_en AT TIME ZONE 'America/Montevideo')::date
              FROM pagos pg JOIN prestamos pr ON pr.id = pg.prestamo_id
             WHERE pr.disapp_credit_ref = %s AND pg.anulado = false
               AND pg.disapp_pago_id IS NOT NULL AND abs(pg.monto - %s) < %s
             ORDER BY pg.registrado_en DESC
        """, (ref, dif, TOL))
        cand = cur.fetchall()
        if len(cand) >= 1:
            x = cand[0]
            anular.append({"pago_id": str(x[0]), "ref": ref, "cliente": nom,
                           "monto": float(x[1]), "disapp_pago_id": str(x[2]),
                           "fecha": str(x[3]), "dif": dif})
        else:
            sin_match.append((nom, ref, dif))

    print(f"{'='*100}\n  ALINEAR CON DISAPP: cobros revertidos allá   "
          f"{'🔴 ANULANDO' if COMMIT else '🟡 SOLO MIRANDO'}\n{'='*100}")
    print(f"\n  ✗ SE ANULAN (la diferencia es exactamente ese pago): {len(anular)}   "
          f"{money(sum(a['monto'] for a in anular))}")
    for a in sorted(anular, key=lambda x: -x["monto"]):
        print(f"     {a['cliente'][:28]:28s} {a['ref']:16s} {money(a['monto']):>11s}  "
              f"del {a['fecha']}  (id Disapp {a['disapp_pago_id']})")
    print(f"\n  ⏸️  se dejan (residuo de varios pagos, no se puede saber cuál): {len(sin_match)}")
    for nom, ref, d in sorted(sin_match, key=lambda x: -x[2]):
        print(f"     {str(nom)[:28]:28s} {ref:16s} sobra {money(d):>10s}")

    if not COMMIT:
        print("\n🟡 No se tocó nada. Aplicar con --commit.\n")
        return
    if not anular:
        print("\nNada que alinear.\n")
        return

    cur.execute("SELECT id FROM usuarios WHERE rol='admin' AND nombre ILIKE 'Mauricio%' LIMIT 1")
    fila = cur.fetchone()
    if not fila:
        cur.execute("SELECT id FROM usuarios WHERE rol='admin' ORDER BY nombre LIMIT 1")
        fila = cur.fetchone()
    if not fila:
        print("🔴 No hay admin para firmar la anulación. Abortado.")
        return

    log = os.path.join(RAIZ, "scripts", f"_revert_anulados_disapp_{dt.date.today():%Y%m%d}.json")
    with open(log, "w", encoding="utf-8") as fh:
        json.dump(anular, fh, ensure_ascii=False, indent=1, default=str)

    cur.execute("""
        UPDATE pagos
           SET anulado = true, anulado_en = now(), anulado_por = %s, motivo_anulacion = %s
         WHERE id = ANY(%s) AND anulado = false
    """, (fila[0], MOTIVO, [a["pago_id"] for a in anular]))
    cn.commit()
    print(f"\n  ✓ anulados: {cur.rowcount}    log de reversa: {log}")
    cur.close()
    cn.close()


if __name__ == "__main__":
    main()
