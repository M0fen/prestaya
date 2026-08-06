# -*- coding: utf-8 -*-
"""
¿Por qué tenemos 2.860 activos y el export de Disapp 2.815?

Desglosa la diferencia crédito por crédito para saber si sobra algo que no
debería estar vivo (y que mañana el cobrador saldría a cobrar) o si cada
excedente tiene una razón conocida. Además lista los activos SOBRE-COBRADOS
(pagado > total), que es plata cobrada de más a un cliente.

Solo LECTURA.
"""
import os
import re
import ssl
from collections import Counter
from urllib.parse import unquote

import openpyxl
import pg8000.dbapi

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORT = r"C:\Users\Carlos\migracion\creditos_2026-08-06_04-07.xlsx"


def conectar():
    with open(os.path.join(RAIZ, ".env.local"), encoding="utf-8") as fh:
        url = next(l.split("=", 1)[1].strip().strip('"').strip("'")
                   for l in fh if l.startswith("SUPABASE_DB_URL="))
    m = re.match(r"postgres(?:ql)?://([^:]+):([^@]+)@([^:/]+):(\d+)/(.+)", url)
    usr, pw, host, port, base = m.groups()
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return pg8000.dbapi.connect(user=unquote(usr), password=unquote(pw), host=host,
                                port=int(port), database=base.split("?")[0], ssl_context=ctx)


def money(n) -> str:
    return "$ " + f"{round(float(n or 0)):,}".replace(",", ".")


def main() -> None:
    wb = openpyxl.load_workbook(EXPORT, read_only=True, data_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    iRef, iEst = hdr.index("Crédito #"), hdr.index("Estado")
    refs_export, estados = set(), Counter()
    for r in ws.iter_rows(min_row=2, values_only=True):
        ref = str(r[iRef] or "").strip()
        if ref:
            refs_export.add(ref)
            estados[str(r[iEst] or "")] += 1
    wb.close()
    print(f"Export: {len(refs_export)} créditos · estados: {dict(estados)}")

    cn = conectar()
    cur = cn.cursor()
    cur.execute("""
        SELECT p.id, p.disapp_credit_ref, p.disapp_credit_id, p.cuota_diaria, p.total_dias,
               p.pagado_acum, p.fecha_inicio, cl.nombre, u.nombre, z.nombre, p.origen
          FROM prestamos p
          JOIN clientes cl ON cl.id = p.cliente_id
          LEFT JOIN usuarios u ON u.id = p.cobrador_id
          LEFT JOIN zonas z ON z.id = u.zona_id
         WHERE p.estado = 'activo'
    """)
    filas = cur.fetchall()
    print(f"Base: {len(filas)} activos\n")

    nativos, ausentes, ok = [], [], 0
    for f in filas:
        ref = f[1]
        if not ref:
            nativos.append(f)          # nacido en Presta Ya (no tiene ref de Disapp)
        elif ref in refs_export:
            ok += 1
        else:
            ausentes.append(f)

    print("═══ DESGLOSE DE LOS ACTIVOS ═══")
    print(f"  ✅ están también activos en el export:      {ok}")
    print(f"  🆕 nativos de Presta Ya (sin ref Disapp):   {len(nativos)}")
    print(f"  ⚠️  con ref pero AUSENTES del export:        {len(ausentes)}")

    if nativos:
        print("\n  Nativos (creados en la app — es correcto que existan):")
        for f in nativos[:12]:
            print(f"     {str(f[7])[:30]:30s} {str(f[9] or '-')[:14]:14s} cuota {money(f[3])} × {f[4]}  inicio {f[6]}")
        if len(nativos) > 12:
            print(f"     … y {len(nativos)-12} más")

    if ausentes:
        tot = sum(max(0.0, float(f[3]) * int(f[4]) - float(f[5] or 0)) for f in ausentes)
        print(f"\n  ⚠️ AUSENTES del export pero VIVOS acá — saldo {money(tot)}:")
        print("     (si Disapp los cerró, mañana el cobrador sale a cobrar algo que ya no existe)")
        for f in sorted(ausentes, key=lambda x: -(float(x[3]) * int(x[4]) - float(x[5] or 0)))[:20]:
            saldo = max(0.0, float(f[3]) * int(f[4]) - float(f[5] or 0))
            print(f"     {str(f[7])[:28]:28s} {str(f[1]):16s} {str(f[9] or '-')[:12]:12s} saldo {money(saldo):>12s}")
        if len(ausentes) > 20:
            print(f"     … y {len(ausentes)-20} más")

    # Sobre-cobrados: pagado por encima del total del crédito.
    cur.execute("""
        SELECT cl.nombre, p.disapp_credit_ref, p.cuota_diaria * p.total_dias, p.pagado_acum,
               u.nombre
          FROM prestamos p JOIN clientes cl ON cl.id = p.cliente_id
          LEFT JOIN usuarios u ON u.id = p.cobrador_id
         WHERE p.estado='activo' AND p.pagado_acum > p.cuota_diaria * p.total_dias + 0.5
         ORDER BY p.pagado_acum - p.cuota_diaria * p.total_dias DESC
    """)
    sobre = cur.fetchall()
    print(f"\n═══ ACTIVOS SOBRE-COBRADOS: {len(sobre)} ═══")
    if sobre:
        print("  (el cliente pagó MÁS que el total del crédito: hay que devolverle o cerrarlo)")
        for s in sobre[:20]:
            exceso = float(s[3]) - float(s[2])
            print(f"     {str(s[0])[:28]:28s} {str(s[1] or '-'):16s} total {money(s[2]):>12s}  pagado {money(s[3]):>12s}  sobra {money(exceso):>10s}")

    cur.close()
    cn.close()


if __name__ == "__main__":
    main()
