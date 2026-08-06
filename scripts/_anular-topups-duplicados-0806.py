# -*- coding: utf-8 -*-
"""
ANULA los ajustes sintéticos de reconciliación que quedaron DUPLICADOS por el
pago real que llegó después.

Qué pasó: cuando un crédito quedaba por debajo del 'Pagos' de Disapp, el empalme
sembraba un ajuste exacto (`origen = 'reconciliacion_*'`) para cuadrarlo. En la
corrida siguiente, el recaudo REAL de ese mismo día llegó en el export nuevo y,
por ser posterior al corte, entró SIN cap. Resultado: la misma plata contada dos
veces — el cliente figura pagando de más y su saldo queda por debajo de lo real.

Regla para anular (conservadora, solo con evidencia):
  · existe un ajuste sintético y un pago real de Disapp para el MISMO crédito,
    el MISMO día y el MISMO monto, y
  · anular el ajuste ACERCA nuestro `pagado_acum` a la verdad, que es:
      - el 'Pagos' del export de Disapp, si el crédito está ahí;
      - el total del crédito, si ya no está (Disapp lo cerró) y quedó sobre-cobrado.

Los pagos NUNCA se borran: se anulan con quién y por qué (libro inmutable).
El ajuste sintético es el que se anula, jamás el cobro real.

  Dry-run:  python scripts/_anular-topups-duplicados-0806.py
  Aplicar:  python scripts/_anular-topups-duplicados-0806.py --commit
"""
import json
import os
import re
import ssl
import sys
import datetime as dt
from urllib.parse import unquote

import openpyxl
import pg8000.dbapi

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORT = r"C:\Users\Carlos\migracion\creditos_2026-08-06_04-07.xlsx"
COMMIT = "--commit" in sys.argv
TOL = 1.0
MOTIVO = ("Ajuste de reconciliación duplicado por el recaudo real de Disapp del mismo "
          "día (empalme 06-08): la misma plata estaba contada dos veces.")


def num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def conectar():
    with open(os.path.join(RAIZ, ".env.local"), encoding="utf-8") as fh:
        url = next(l.split("=", 1)[1].strip().strip('"').strip("'")
                   for l in fh if l.startswith("SUPABASE_DB_URL="))
    m = re.match(r"postgres(?:ql)?://([^:]+):([^@]+)@([^:/]+):(\d+)/(.+)", url)
    usr, pw, host, port, base = m.groups()
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return pg8000.dbapi.connect(user=unquote(usr), password=unquote(pw), host=host,
                                port=int(port), database=base.split("?")[0], ssl_context=ctx)


def money(n) -> str:
    return "$ " + f"{round(float(n or 0)):,}".replace(",", ".")


def main() -> None:
    wb = openpyxl.load_workbook(EXPORT, read_only=True, data_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    iRef, iPag = hdr.index("Crédito #"), hdr.index("Pagos")
    disapp = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        ref = str(r[iRef] or "").strip()
        if ref:
            disapp[ref] = num(r[iPag])
    wb.close()

    cn = conectar()
    cur = cn.cursor()
    cur.execute("""
        WITH t AS (
          SELECT id, prestamo_id, monto,
                 (registrado_en AT TIME ZONE 'America/Montevideo')::date d
            FROM pagos WHERE anulado = false AND origen LIKE 'reconciliacion%'
        ), r AS (
          SELECT prestamo_id, monto,
                 (registrado_en AT TIME ZONE 'America/Montevideo')::date d
            FROM pagos WHERE anulado = false AND origen = 'disapp_import'
        )
        SELECT DISTINCT t.id, cl.nombre, p.disapp_credit_ref, t.d, t.monto,
               p.cuota_diaria * p.total_dias, p.pagado_acum, p.estado
          FROM t
          JOIN r ON r.prestamo_id = t.prestamo_id AND r.monto = t.monto AND r.d = t.d
          JOIN prestamos p ON p.id = t.prestamo_id
          JOIN clientes cl ON cl.id = p.cliente_id
         ORDER BY t.monto DESC
    """)
    filas = cur.fetchall()

    anular, dejar = [], []
    for pid, nombre, ref, dia, monto, total, pagado, estado in filas:
        monto, total, pagado = float(monto), float(total), float(pagado)
        verdad = disapp.get(ref)
        if verdad is not None:
            # Está en el export: la verdad es su columna 'Pagos'.
            antes, despues = abs(pagado - verdad), abs(pagado - monto - verdad)
            motivo = f"Disapp dice {money(verdad)}, tenemos {money(pagado)}"
        elif pagado > total + TOL:
            # Disapp lo cerró y quedó sobre-cobrado: la verdad es el total.
            antes, despues = abs(pagado - total), abs(pagado - monto - total)
            motivo = f"cerrado en Disapp y sobre-cobrado: total {money(total)}, tenemos {money(pagado)}"
        else:
            dejar.append((nombre, ref, monto, "sin evidencia concluyente"))
            continue
        if despues < antes - TOL:
            anular.append({"pago_id": str(pid), "cliente": nombre, "ref": ref, "dia": str(dia),
                           "monto": monto, "motivo": motivo,
                           "acerca": round(antes - despues, 2)})
        else:
            dejar.append((nombre, ref, monto, f"anularlo NO acerca a la verdad ({motivo})"))

    print(f"{'='*100}\n  AJUSTES SINTÉTICOS DUPLICADOS   {'🔴 APLICANDO' if COMMIT else '🟡 DRY-RUN'}\n{'='*100}")
    print(f"  candidatos (ajuste + pago real mismo crédito/día/monto): {len(filas)}")
    print(f"  ✅ SE ANULAN (acercan a la verdad):  {len(anular)}   {money(sum(a['monto'] for a in anular))}")
    print(f"  ⏸️  se dejan (sin evidencia):          {len(dejar)}")
    for a in anular[:20]:
        print(f"     {a['cliente'][:28]:28s} {a['ref']:16s} {a['dia']}  {money(a['monto']):>10s}  — {a['motivo']}")
    if len(anular) > 20:
        print(f"     … y {len(anular)-20} más")
    if dejar[:8]:
        print("\n  Se dejan como están:")
        for n, ref, m, por in dejar[:8]:
            print(f"     {str(n)[:28]:28s} {str(ref):16s} {money(m):>10s}  — {por}")

    if not COMMIT:
        print("\n🟡 DRY-RUN: no se escribió nada. Aplicar con --commit.\n")
        return
    if not anular:
        print("\nNada que anular.\n")
        return

    log = os.path.join(RAIZ, "scripts", f"_revert_topups_{dt.date.today():%Y%m%d}.json")
    with open(log, "w", encoding="utf-8") as fh:
        json.dump(anular, fh, ensure_ascii=False, indent=1)

    # La base exige registrar QUIÉN anula (chk_pago_anulacion_completa): el libro
    # de pagos no admite una anulación anónima. Se firma con el admin dueño.
    cur.execute("SELECT id FROM usuarios WHERE rol='admin' AND nombre ILIKE 'Mauricio%' LIMIT 1")
    fila = cur.fetchone()
    if not fila:
        cur.execute("SELECT id FROM usuarios WHERE rol='admin' ORDER BY nombre LIMIT 1")
        fila = cur.fetchone()
    if not fila:
        print("\n🔴 No hay usuario admin para firmar la anulación. Abortado.")
        return
    quien = fila[0]

    ids = [a["pago_id"] for a in anular]
    cur.execute("""
        UPDATE pagos
           SET anulado = true, anulado_en = now(), anulado_por = %s, motivo_anulacion = %s
         WHERE id = ANY(%s) AND anulado = false
    """, (quien, MOTIVO, ids))
    cn.commit()
    print(f"\n  ✓ anulados: {cur.rowcount}")
    print(f"  log de reversa: {log}")

    cur.execute("""
        SELECT count(*) FROM prestamos
         WHERE estado='activo' AND pagado_acum > cuota_diaria*total_dias + 0.5
    """)
    print(f"  activos sobre-cobrados que quedan: {cur.fetchone()[0]}")
    cur.close()
    cn.close()


if __name__ == "__main__":
    main()
