#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EMPALME TOTAL 2026-08-04 — sincronizar Presta Ya con Disapp a día de hoy (todas
las zonas), la noche previa al arranque del piloto.

Qué hace (en orden, cada paso idempotente — se puede RE-CORRER con exports más
nuevos, p.ej. mañana con los recaudos del 08-04):
  1. CLIENTES nuevos de Disapp que no tenemos (por disapp_id; documento choca → NULL).
  2. CRÉDITOS activos de Disapp que no tenemos (por disapp_credit_id y ref).
  3. ASIGNACIONES: cada crédito nuevo deja al cliente en la ruta de su cobrador
     (INV10: crédito activo sin ruta = crítico). No se pisa la ruta de terceros
     con crédito activo propio.
  4. RECAUDOS: inserta los pagos frescos (≥ 2026-07-21) deduplicados por
     disapp_pago_id, CAPADOS POR CRÉDITO contra la columna 'Pagos' del export de
     créditos de HOY. El cap tira desde el FRENTE (los más viejos): el excedente
     típico es el recaudo del 07-21 que YA vive dentro del ajuste de la
     sincronización de Zona Centro — insertarlo lo contaría dos veces.
     Para créditos nuevos entra TODA su historia de recaudos (con el mismo cap).
  5. TOP-UPS: si tras el paso 4 un crédito con target sigue por debajo del
     'Pagos' de Disapp (p.ej. semanales de supervisor, cuyos pagos Disapp no
     exporta), se siembra UN ajuste exacto (disapp_pago_id = recon-<hoy>-<id>).
  6. FINALIZAR: activos nuestros que ya NO están activos en Disapp y quedaron
     con saldo (refinanciados/cerrados allá) → estado 'finalizado' +
     finalizado_en. EXCEPCIÓN: clientes BORRADOS en Disapp se dejan activos
     (decisión: acá no hay clientes fantasma; se siguen cobrando).

La VERDAD por crédito es la columna 'Pagos' del export de créditos (identidad
Pagos+Saldo==Total c/Intereses verificada en el 100% de los activos). El estado
del cartón se deriva (FIFO); acá solo se alinea el libro de pagos.

  Dry-run:  python scripts/empalme-0804.py --env-file .env.local
  Escribir: python scripts/empalme-0804.py --env-file .env.local --commit
  Opcional: --src PATH (default C:\\Users\\Carlos\\migracion) · --forzar (saltea
            la guardia de pagos nativos de la app)

NO borra ni edita nada jamás (0126 lo veta a nivel BD, además). Solo inserta
pagos/clientes/créditos/asignaciones y avanza estados activo→finalizado.
"""
import sys, os, json, datetime as dt, urllib.parse, urllib.request
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import empalme_disapp as E  # parsers + consolidar + get_rows/upsert/http probados

def arg(flag, default=None):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else default

COMMIT = "--commit" in sys.argv
FORZAR = "--forzar" in sys.argv
SALTEAR = "--saltear-choques" in sys.argv  # saltea SOLO las filas que chocan (la app manda) y sigue
# ⚠️ Levanta el CANDADO del paso 6: cierra TAMBIÉN los créditos ausentes del
# export que todavía deben plata. Existe para la RECARGA TOTAL: cuando el export
# es completo y confiable, la ausencia SÍ significa baja, y sin esto el espejo no
# converge nunca (arrastraríamos créditos que Disapp ya cerró y nuestros totales
# quedarían siempre por encima). Nunca se activa solo: el default es no cerrar.
# La lista completa queda SIEMPRE en CSV, aunque no se use el flag, para que un
# humano la revise ANTES de decidir.
CERRAR_CON_SALDO = "--cerrar-con-saldo" in sys.argv
SRC = arg("--src", r"C:\Users\Carlos\migracion")
ENVF = arg("--env-file", ".env.local")
FRONTERA = dt.date(2026, 7, 21)   # el último import de recaudos llegó hasta el 07-20
# CORTE del export de CRÉDITOS: hasta qué día (inclusive) su columna 'Pagos'
# refleja los recaudos. El CAP anti-duplicado solo aplica a recaudos ≤ corte;
# los posteriores entran SIEMPRE (el target viejo no los conoce y caparlos
# tiraría cobros reales). Si mañana llega solo el excel de recaudos del 08-04
# sin créditos frescos: dejar --corte 2026-08-03. Si llegan créditos frescos:
# subir el corte a la fecha de ese export.
CORTE = dt.date.fromisoformat(arg("--corte", "2026-08-03"))
TOL = 1.0
HOY = dt.date.today()
SELLO = HOY.strftime("%Y%m%d")

env = E.load_env(ENVF)
url = (env.get("SUPABASE_URL") or env.get("NEXT_PUBLIC_SUPABASE_URL") or "").rstrip("/")
key = env.get("SUPABASE_SERVICE_ROLE_KEY")
if not url or not key:
    sys.exit(f"Faltan SUPABASE_URL/SERVICE_ROLE_KEY en {ENVF}")
db = {"url": url, "key": key, "host": urllib.parse.urlparse(url).netloc}
print(f"EMPALME 0804 → {db['host'].split('.')[0]}  | modo: {'🔴 COMMIT (escribe)' if COMMIT else '🟡 DRY-RUN'}")

# ══ 0. Cargar exports + base ════════════════════════════════════════════════
d = E.consolidar(SRC)

# Créditos del export con TODO (incluye Tasa %): re-leo el xlsx de créditos acá
# para no depender de que load_creditos exponga la tasa.
import openpyxl, glob as _glob
creds_hoy = {}
for path in sorted(_glob.glob(os.path.join(SRC, "creditos*.xlsx"))):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdr = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    hm = {h.lower(): i for i, h in enumerate(hdr) if h}
    def gv(r, *names):
        for n in names:
            i = hm.get(n.lower())
            if i is not None:
                return r[i]
        return None
    for r in ws.iter_rows(min_row=2, values_only=True):
        cid = gv(r, "ID Crédito", "ID Credito")
        if cid in (None, ""):
            continue
        cid = str(cid).strip()
        if cid in creds_hoy:
            continue
        if (str(gv(r, "Estado") or "")).strip().lower() != "activo":
            continue
        mod = (str(gv(r, "Modalidad") or "diaria")).strip().lower()
        tasa_raw = gv(r, "Tasa %", "Tasa")
        tasa = E.parse_money_text(str(tasa_raw).replace("%", "")) if tasa_raw not in (None, "") else None
        creds_hoy[cid] = {
            "cid": cid,
            "ref": (str(gv(r, "Crédito #", "Credito #") or "")).strip() or None,
            "id_cliente": (str(gv(r, "ID Cliente") or "")).strip() or None,
            "id_vendedor": (str(gv(r, "ID Vendedor") or "")).strip() or None,
            "vendedor": (str(gv(r, "Vendedor") or "")).strip(),
            "frecuencia": E.MODALIDAD_FREC.get(mod, "diario"),
            "valor": E.parse_money_text(gv(r, "Valor Crédito", "Valor Credito")) or 0,
            "tci": E.parse_money_text(gv(r, "Total c/ Intereses", "Total con Intereses")) or 0,
            "pagos": E.parse_money_text(gv(r, "Pagos")) or 0,
            "cuota": E.parse_money_text(gv(r, "Valor Cuota")) or 0,
            "cuotas": int(gv(r, "Cuotas") or 0),
            "fecha": E.parse_date(gv(r, "Fecha Crédito", "Fecha Credito")),
            "tasa": round(tasa) if tasa is not None else None,
        }
    wb.close()
ref2hoy = {c["ref"]: c for c in creds_hoy.values() if c["ref"]}
print(f"  export créditos HOY: {len(creds_hoy)} activos · export clientes: {len(d['clientes'])} · pagos únicos: {len(d['pagos'])}")

# INTERLOCK del corte: si el export de créditos es MÁS NUEVO que --corte, su
# columna 'Pagos' ya contiene recaudos posteriores al corte → el top-up los
# duplicaría (fila real + ajuste). Se corta acá, no en producción.
max_fc = max((c["fecha"] for c in creds_hoy.values() if c["fecha"]), default=None)
if max_fc and max_fc > CORTE:
    sys.exit(
        f"🔴 ABORTA: el export de créditos trae 'Fecha Crédito' hasta {max_fc} > corte {CORTE}.\n"
        f"   Su 'Pagos' ya refleja días posteriores al corte: volvé a correr con --corte {max_fc}\n"
        f"   (la fecha real del export). Con el corte viejo, cada recaudo nuevo entraría DOS veces."
    )

usuarios = E.get_rows(db, "usuarios", "id,nombre,rol,zona_id,disapp_vendedor_id,activo")
zonas = E.get_rows(db, "zonas", "id,nombre")
zona_nom = {z["id"]: z["nombre"] for z in zonas}
vend2user = {str(u["disapp_vendedor_id"]): u for u in usuarios if u.get("disapp_vendedor_id") is not None}
user_zona = {u["id"]: (zona_nom.get(u["zona_id"]) if u.get("zona_id") else None) or "(sin zona)" for u in usuarios}

clientes_db = E.get_rows(db, "clientes", "id,disapp_id,documento,activo")
cliDe = {str(c["disapp_id"]): c["id"] for c in clientes_db if c.get("disapp_id")}
docs_db = {c["documento"] for c in clientes_db if c.get("documento")}
ids_cli_file = set(d["clientes"].keys())
borrados_en_disapp = {c["id"] for c in clientes_db if c.get("disapp_id") and c["disapp_id"] not in ids_cli_file}

pres = E.get_rows(db, "prestamos", "id,cliente_id,cobrador_id,disapp_credit_id,disapp_credit_ref,estado,cuota_diaria,total_dias,pagado_acum,finalizado_en")
by_ref_db = {p["disapp_credit_ref"]: p for p in pres if p.get("disapp_credit_ref")}
ids_credit_db = {str(p["disapp_credit_id"]) for p in pres if p.get("disapp_credit_id")}
activos_db = [p for p in pres if p["estado"] == "activo"]

# pagos ya en la base (para no contar dos veces en el CAP al re-correr)
ya_ids = set()
for row in E.get_rows(db, "pagos", "id,disapp_pago_id", {"registrado_en": "gte.2026-06-01"}):
    if row.get("disapp_pago_id"):
        ya_ids.add(str(row["disapp_pago_id"]))
print(f"  base: {len(clientes_db)} clientes · {len(pres)} créditos ({len(activos_db)} activos) · pagos recientes ya importados: {len(ya_ids)}")

# ══ LA APP MANDA ═══════════════════════════════════════════════════════════
# Desde que un cobrador registra pagos NATIVOS (origen NULL) en la app, su
# cartera es app-autoritativa: el export de Disapp para su zona queda VIEJO
# (ZC deja de cargar en Disapp desde el 08-05). A esos créditos NO se les
# resucita estado, NO se les siembra top-up, NO se los finaliza por ausencia
# en el export, y sus candidatos con fecha ≥ piloto se descartan (re-capturas).
PILOTO_DESDE = dt.date(2026, 8, 5)
_nativos_piloto = E.get_rows(db, "pagos", "id,registrado_por,prestamo_id,origen,registrado_en",
                             {"anulado": "eq.false", "registrado_en": f"gte.{PILOTO_DESDE}T00:00:00-03:00"})
_nativos_piloto = [n for n in _nativos_piloto if n.get("origen") is None]
cobradores_vivos = {n["registrado_por"] for n in _nativos_piloto if n.get("registrado_por")}
prestamos_nativos = {n["prestamo_id"] for n in _nativos_piloto}


def _dia_uy(iso):
    """Día calendario uruguayo (UTC−3) de un timestamptz ISO."""
    if not iso:
        return None
    s = str(iso).replace("Z", "+00:00")
    try:
        t = dt.datetime.fromisoformat(s)
    except ValueError:
        return None
    if t.tzinfo is None:
        t = t.replace(tzinfo=dt.timezone.utc)
    return (t.astimezone(dt.timezone(dt.timedelta(hours=-3)))).date()


# (crédito, día UY) que la app YA cobró de verdad. Es la llave fina para decidir
# si una fila del export es una re-captura o un cobro que solo vive en Disapp.
nativo_dia = {(n["prestamo_id"], _dia_uy(n.get("registrado_en"))) for n in _nativos_piloto}
nativo_dia.discard((None, None))
print(f"  app-autoritativo: {len(cobradores_vivos)} cobradores con pagos nativos desde {PILOTO_DESDE} ({len(_nativos_piloto)} pagos)")

# Créditos con imports ANULADOS en la app: el admin corrigió un asiento de
# Disapp a propósito → el target de Disapp ya no es la verdad para ese crédito;
# sembrar top-up lo DES-anularía cada noche. Se saltean y se listan.
_anul_imp = E.get_rows(db, "pagos", "id,prestamo_id,disapp_pago_id", {"anulado": "eq.true", "disapp_pago_id": "not.is.null"})
prestamos_con_import_anulado = {a["prestamo_id"] for a in _anul_imp}
_id2ref = {p["id"]: p.get("disapp_credit_ref") for p in pres}
refs_con_import_anulado = {_id2ref.get(pid) for pid in prestamos_con_import_anulado} - {None}
if refs_con_import_anulado:
    print(f"  créditos con imports anulados en la app (sin top-up, a revisión): {len(refs_con_import_anulado)}")

# ══ 1. Clientes nuevos ══════════════════════════════════════════════════════
faltan_creds = [c for c in creds_hoy.values() if c["ref"] not in by_ref_db and c["cid"] not in ids_credit_db]
cli_nuevos = {}
docs_previstos = set()
for c in faltan_creds:
    did = c["id_cliente"]
    if not did or did in cliDe or did in cli_nuevos:
        continue
    f = d["clientes"].get(did) or {}
    doc = f.get("documento")
    if doc and (doc in docs_db or doc in docs_previstos):
        doc = None  # misma regla del empalme: el 1ro conserva el documento
    if doc:
        docs_previstos.add(doc)
    cli_nuevos[did] = {
        "disapp_id": did,
        "nombre": f.get("nombre") or "(sin nombre)",
        "documento": doc,
        "telefono": f.get("telefono"),
        "direccion": f.get("direccion"),
        "activo": True,
        "origen": "oficina",
    }

# ══ 2. Créditos nuevos ══════════════════════════════════════════════════════
sin_vendedor = [c for c in faltan_creds if str(c["id_vendedor"]) not in vend2user]
crear = [c for c in faltan_creds if str(c["id_vendedor"]) in vend2user]

# ══ 2b. RESUCITAR: activos HOY en Disapp que acá quedaron finalizados ═══════
# La sincronización ZC del 07-21 finalizó 245 créditos porque el export de aquel
# día no los listaba; el export de HOY dice que siguen activos (con saldo). La
# verdad es el export más fresco → se reactivan (0126 lo permite vía service).
resucitar, resucitar_saltados = [], []
def _dia_de(ts):
    return str(ts)[:10] if ts else None
for ref, c in ref2hoy.items():
    p = by_ref_db.get(ref)
    if not p or p["estado"] == "activo":
        continue
    # (a) cartera app-autoritativa: una renovación/finalización hecha EN LA APP
    #     no se revierte porque un export viejo todavía liste el ref como Activo.
    if p["cobrador_id"] in cobradores_vivos or p["id"] in prestamos_nativos:
        resucitar_saltados.append((ref, "zona viva (la app manda)")); continue
    # (b) finalización de la app POSTERIOR al corte del export: la app es más nueva.
    fe = _dia_de(p.get("finalizado_en"))
    if fe and fe > CORTE.isoformat():
        resucitar_saltados.append((ref, f"finalizado en la app el {fe} > corte")); continue
    # (c) términos deben coincidir con el export (¿ref reciclado por Disapp?).
    if round(float(p["cuota_diaria"] or 0)) != round(c["cuota"] or 0) or int(p["total_dias"] or 0) != int(c["cuotas"] or 0):
        resucitar_saltados.append((ref, "términos distintos (¿ref reciclado?)")); continue
    resucitar.append({"id": p["id"], "ref": ref, "estado": p["estado"],
                      "cliente_id": p["cliente_id"], "cobrador_id": p["cobrador_id"]})

# ══ 4a. Candidatos de pagos por crédito (antes de asignaciones p/ conocer scope)
HOY_UY = dt.datetime.now(dt.timezone(dt.timedelta(hours=-3))).date()
refs_nuevos = {c["ref"] for c in crear}
# Créditos existentes con pagado 0 y ref en el export: quedaron a medias por un
# corte previo (crédito creado, pagos nunca imputados) → se les da la historia
# COMPLETA como a un nuevo, no solo lo fresco.
refs_cero = {r for r, p in by_ref_db.items() if round(float(p["pagado_acum"] or 0)) == 0}
cand_por_ref = defaultdict(list)
clamp_hoy = recapturas = 0
for p in d["pagos"].values():
    if not p["ref"] or not p["fecha"] or str(p["id_pago"]) in ya_ids:
        continue
    # Nunca importar recaudos del DÍA EN CURSO: contaminan la custodia de HOY
    # (esperado del cierre, float, push). El ritual siempre importa AYER.
    if p["fecha"] >= HOY_UY:
        clamp_hoy += 1
        continue
    pdb = by_ref_db.get(p["ref"])
    # ⚠️ RE-CAPTURA = el MISMO crédito cobrado el MISMO día en los dos lados.
    # La regla vieja descartaba por COBRADOR: si el cobrador había usado la app
    # aunque fuera una vez, se tiraba TODO lo que Disapp trajera de su cartera.
    # Medido contra el export del 08-06 (scripts/_verificar-recapturas-0806.py):
    # de 300 filas descartadas, solo 84 tenían de verdad un pago nativo ese día —
    # las otras 216 ($60.562, 9 cobradores) existían SOLO en Disapp. Tirarlas
    # significa que al día siguiente la app le vuelve a pedir esa cuota a 216
    # clientes que YA pagaron: el doble cobro del día 1, otra vez. Ahora se exige
    # la coincidencia fina (crédito, día); si la app no tiene ese cobro, ENTRA.
    if pdb and p["fecha"] >= PILOTO_DESDE and (pdb["id"], p["fecha"]) in nativo_dia:
        recapturas += 1
        continue
    if pdb and (p["fecha"] >= FRONTERA or p["ref"] in refs_cero):
        cand_por_ref[p["ref"]].append(p)      # existente: lo fresco (o TODO si quedó en cero)
    elif p["ref"] in refs_nuevos:
        cand_por_ref[p["ref"]].append(p)      # crédito nuevo: toda su historia
for lst in cand_por_ref.values():
    lst.sort(key=lambda p: (p["fecha"], str(p["id_pago"])))

# CAP por crédito con target (matcheados y nuevos): tirar desde el FRENTE.
insertar = []            # pagos reales a insertar
descartes = []           # (ref, pago) descartados por el cap
topups = []              # ajustes exactos
sim_final = {}           # ref -> pagado final simulado
refs_crear = {c["ref"]: c for c in crear if c["ref"]}
sim_corte = {}  # ref -> pagado simulado SOLO hasta el corte (base p/ top-ups)
for ref, lst in cand_por_ref.items():
    hoy_file = ref2hoy.get(ref)
    if ref in by_ref_db:
        base_pagado = round(float(by_ref_db[ref]["pagado_acum"] or 0), 2)
    else:
        base_pagado = 0.0
    capables = [p for p in lst if p["fecha"] <= CORTE]   # el target los conoce
    post = [p for p in lst if p["fecha"] > CORTE]        # posteriores: entran siempre
    if hoy_file:  # hay target
        target = round(hoy_file["pagos"], 2)
        exceso = round(base_pagado + sum(p["monto"] or 0 for p in capables) - target, 2)
        keep = list(capables)
        while exceso > TOL and keep:
            m = keep[0]["monto"] or 0
            if m <= exceso + TOL:
                descartes.append((ref, keep.pop(0)))
                exceso = round(exceso - m, 2)
            else:
                break  # el próximo no entra entero en el exceso: no tirar de más
        insertar.extend((ref, p) for p in keep + post)
        sim_corte[ref] = round(base_pagado + sum(p["monto"] or 0 for p in keep), 2)
        sim_final[ref] = round(sim_corte[ref] + sum(p["monto"] or 0 for p in post), 2)
    else:  # stale (activo acá, cerrado allá): sin target, entra todo lo fresco
        insertar.extend((ref, p) for p in lst)
        sim_final[ref] = round(base_pagado + sum(p["monto"] or 0 for p in lst), 2)

# ══ 5. Top-ups exactos p/ créditos con target que quedan cortos ═════════════
topups_saltados = []
for ref, c in ref2hoy.items():
    en_db = ref in by_ref_db
    es_nuevo = ref in refs_crear
    if not en_db and not es_nuevo:
        continue
    pdb = by_ref_db.get(ref)
    # La app manda: crédito con actividad nativa no recibe ajustes de Disapp.
    if pdb and (pdb["cobrador_id"] in cobradores_vivos or pdb["id"] in prestamos_nativos):
        continue
    # El admin anuló un import a propósito: el target de Disapp dejó de ser la
    # verdad para este crédito — un top-up des-anularía esa decisión cada noche.
    if ref in refs_con_import_anulado:
        topups_saltados.append(ref)
        continue
    base_pagado = round(float(pdb["pagado_acum"] or 0), 2) if pdb else 0.0
    # El hueco se mide contra lo simulado HASTA EL CORTE: el target no conoce
    # los recaudos posteriores, y esos no tapan un faltante de la ventana vieja.
    fin_corte = sim_corte.get(ref, sim_final.get(ref, base_pagado))
    delta = round(c["pagos"] - fin_corte, 2)
    if delta > TOL:
        topups.append({"ref": ref, "cid": c["cid"], "delta": delta, "vend": c["vendedor"]})
        sim_final[ref] = round(sim_final.get(ref, base_pagado) + delta, 2)

# ══ 6. Finalizar stales con saldo (excepto clientes borrados en Disapp) ═════
finalizar = []
mantener_borrados = []
no_finalizar_con_saldo = []  # ausentes del export PERO con deuda viva → nunca se cierran
for p in activos_db:
    ref = p.get("disapp_credit_ref")
    if not ref or ref in ref2hoy:
        continue
    # La app manda: si el cobrador ya opera en la app, que su crédito no esté
    # en el export de Disapp NO significa que se cerró — significa que la zona
    # migró (o que el export vino sin esa zona). JAMÁS finalizar por ausencia.
    if p["cobrador_id"] in cobradores_vivos or p["id"] in prestamos_nativos:
        continue
    total = round(float(p["cuota_diaria"] or 0)) * int(p["total_dias"] or 0)
    fin = sim_final.get(ref, round(float(p["pagado_acum"] or 0), 2))
    if p["cliente_id"] in borrados_en_disapp:
        mantener_borrados.append(p)  # Disapp lo borró; acá se sigue cobrando
        continue
    # ⚠️ CANDADO DURO (08-05): un crédito con SALDO VIVO no se cierra JAMÁS por
    # ausencia en un export. Cerrar deuda es una decisión de negocio (condonar,
    # castigar), nunca el efecto colateral de un archivo que no lo trae.
    # Por qué existe: la guardia de "zona viva" de arriba se apoya en
    # `cobradores_vivos`, que se deriva de pagos nativos desde PILOTO_DESDE. El
    # 08-04 esa fecha era MAÑANA → el set estaba vacío → la guardia fue INERTE y
    # una corrida cerró 370 créditos con $1.382.159 de deuda viva; 84 clientes del
    # piloto desaparecieron de la ruta de su cobrador al día siguiente. Este
    # candado no depende de ninguna fecha ni de quién cobró: si se debe plata, no
    # se toca y se reporta para que lo mire un humano.
    colgado = max(0, round(total - fin))
    if colgado >= 1:
        no_finalizar_con_saldo.append({"id": p["id"], "ref": ref, "colgado": colgado,
                                       "cobrador_id": p["cobrador_id"], "cliente_id": p["cliente_id"]})
        # Con --cerrar-con-saldo (recarga total, export completo) SÍ se cierran:
        # la decisión es humana y explícita, y la lista quedó en el CSV de arriba.
        if not CERRAR_CON_SALDO:
            continue
        finalizar.append({"id": p["id"], "ref": ref, "colgado": colgado})
        continue
    # SALDADO y ausente del export → FINALIZAR TAMBIÉN (auditoría 08-05). El
    # `continue` que había acá ("el cartón derivado ya lo refleja") dejó 678
    # créditos muertos como `activo` para siempre: la ficha mostraba "N créditos
    # activos" en ~915 clientes (Disapp real: 456), mora contaba vencidos de $0,
    # Renovar los ofrecía como candidatos y el panel perdió credibilidad en la
    # presentación. Un crédito que Disapp ya no lista y no debe un peso está
    # TERMINADO: colgado=0 (nada que perseguir), pero el estado debe decirlo.
    finalizar.append({"id": p["id"], "ref": ref, "colgado": max(0, round(total - fin))})

# ══ 3. Asignaciones para créditos nuevos ════════════════════════════════════
asig = E.get_rows(db, "asignaciones", "id,cobrador_id,cliente_id,activo")
act_pair = {(a["cliente_id"], a["cobrador_id"]): a for a in asig if a["activo"]}
inact_pair = {(a["cliente_id"], a["cobrador_id"]): a for a in asig if not a["activo"]}
# créditos activos por (cliente, cobrador) en el estado FINAL (activos − finalizar + nuevos)
finalizar_ids = {f["id"] for f in finalizar}
cred_act_final = defaultdict(int)
for p in activos_db:
    if p["id"] not in finalizar_ids:
        cred_act_final[(p["cliente_id"], p["cobrador_id"])] += 1
asig_activar, asig_crear, asig_bajar = [], [], []
pares_nuevos = set()
for r in resucitar:  # el crédito revive → su cliente vuelve a la ruta del dueño
    par = (r["cliente_id"], r["cobrador_id"])
    cred_act_final[par] += 1
    if par not in act_pair and par not in pares_nuevos:
        pares_nuevos.add(par)
        if par in inact_pair:
            asig_activar.append(inact_pair[par] | {"cliente_id": r["cliente_id"]})
        else:
            asig_crear.append({"cliente_id": r["cliente_id"], "cobrador_id": r["cobrador_id"]})
for c in crear:
    cliente_id = cliDe.get(c["id_cliente"])  # None si el cliente es nuevo (se resuelve al aplicar)
    cobrador_id = vend2user[str(c["id_vendedor"])]["id"]
    key_par = (c["id_cliente"], cobrador_id)
    if key_par in pares_nuevos:
        continue
    pares_nuevos.add(key_par)
    cred_act_final[(cliente_id, cobrador_id)] += 1
    if cliente_id is None:
        asig_crear.append({"cliente_disapp": c["id_cliente"], "cobrador_id": cobrador_id})
        continue
    if (cliente_id, cobrador_id) in act_pair:
        pass  # ya está en la ruta correcta
    elif (cliente_id, cobrador_id) in inact_pair:
        asig_activar.append(inact_pair[(cliente_id, cobrador_id)] | {"cliente_id": cliente_id})
    else:
        asig_crear.append({"cliente_id": cliente_id, "cobrador_id": cobrador_id})
# RED DE SEGURIDAD (crash a mitad de una corrida previa): TODO crédito activo
# matcheado con el export debe tener su par (cliente, cobrador) en ruta — no
# solo los que este plan crea/resucita. Si una corrida anterior murió entre
# crear créditos y asignaciones, esta pasada lo repara sola (INV10 en paz).
for p in activos_db:
    ref = p.get("disapp_credit_ref")
    if not ref or ref not in ref2hoy:
        continue
    par = (p["cliente_id"], p["cobrador_id"])
    if par in act_pair or par in pares_nuevos:
        continue
    pares_nuevos.add(par)
    cred_act_final[par] += 1
    if par in inact_pair:
        asig_activar.append(inact_pair[par] | {"cliente_id": p["cliente_id"]})
    else:
        asig_crear.append({"cliente_id": p["cliente_id"], "cobrador_id": p["cobrador_id"]})

# bajar pares activos de OTROS cobradores sin crédito activo final con ese cliente
clientes_tocados = {cliDe.get(c["id_cliente"]) for c in crear} - {None}
for (cli, cob), a in act_pair.items():
    if cli in clientes_tocados and cred_act_final.get((cli, cob), 0) == 0:
        asig_bajar.append(a)

# ══ Guardia: pagos nativos de la app en la ventana ══════════════════════════
# ⚠️ Nativo = origen IS NULL, filtrado EN PYTHON: un not.in/neq de PostgREST
# EXCLUYE los NULL (SQL trivalente) → la guardia quedaba CIEGA justo a los pagos
# que debía proteger. Hallazgo de la auditoría 08-04: la versión anterior
# imprimía "✓ sin choques (0 revisados)" vacuamente.
nativos = E.get_rows(db, "pagos", "id,prestamo_id,registrado_en,origen,disapp_pago_id",
                     {"anulado": "eq.false", "registrado_en": f"gte.{FRONTERA}"})
nativos = [n for n in nativos if n.get("origen") is None and not str(n.get("disapp_pago_id") or "").startswith("recon-")]
UY = dt.timezone(dt.timedelta(hours=-3))
def dia_uy(ts):
    try:
        return dt.datetime.fromisoformat(str(ts).replace("Z", "+00:00")).astimezone(UY).date().isoformat()
    except ValueError:
        return str(ts)[:10]
nativo_en = {(n["prestamo_id"], dia_uy(n["registrado_en"])) for n in nativos}
choques = []
for ref, p in insertar:
    pdb = by_ref_db.get(ref)
    if pdb and (pdb["id"], p["fecha"].isoformat()) in nativo_en:
        choques.append((ref, p))

# ══ RESUMEN (dry-run y commit) ══════════════════════════════════════════════
def zona_de(c):
    u = vend2user.get(str(c.get("id_vendedor") or ""))
    return user_zona.get(u["id"], "(sin zona)") if u else "(sin usuario)"

print("\n═══ PLAN ═══")
print(f"  1) Clientes a crear: {len(cli_nuevos)}")
pz = defaultdict(lambda: [0, 0.0])
for c in crear:
    pz[zona_de(c)][0] += 1
    pz[zona_de(c)][1] += c["tci"] - c["pagos"]
print(f"  2) Créditos a crear: {len(crear)}  (saldo Disapp ${round(sum(c['tci']-c['pagos'] for c in crear)):,})")
for k, (n, s) in sorted(pz.items(), key=lambda x: -x[1][0]):
    print(f"       {k}: {n}  (saldo ${round(s):,})")
if sin_vendedor:
    print(f"     ⚠ sin vendedor mapeado (NO se crean): {len(sin_vendedor)}")
if resucitar:
    de = defaultdict(int)
    for r in resucitar:
        de[r["estado"]] += 1
    print(f"  2b) Resucitar (activos HOY en Disapp, acá {dict(de)}): {len(resucitar)}")
if resucitar_saltados:
    motivos = defaultdict(int)
    for _, m in resucitar_saltados:
        motivos[m.split(" (")[0]] += 1
    print(f"       NO resucitados (protección): {len(resucitar_saltados)} → {dict(motivos)}")
print(f"  3) Asignaciones: crear {len(asig_crear)} · reactivar {len(asig_activar)} · bajar {len(asig_bajar)}")
n_ins = len(insertar); s_ins = round(sum(p['monto'] or 0 for _, p in insertar))
print(f"  4) Recaudos a insertar: {n_ins}  ${s_ins:,}")
print(f"       descartados por CAP (ya viven en ajustes — no duplicar): {len(descartes)}  ${round(sum(p['monto'] or 0 for _, p in descartes)):,}")
if clamp_hoy:
    print(f"       ⚠ con fecha de HOY o futura (NO entran — el ritual importa AYER): {clamp_hoy}")
if recapturas:
    print(f"       ⚠ re-capturas en Disapp de créditos que la app ya maneja (NO entran): {recapturas}")
print(f"  5) Top-ups exactos: {len(topups)}  ${round(sum(t['delta'] for t in topups)):,}  (fechados al corte {CORTE})")
sup_top = [t for t in topups if "SUPERVISOR" in (t["vend"] or "").upper()]
print(f"       de ellos en créditos de SUPERVISOR: {len(sup_top)}  ${round(sum(t['delta'] for t in sup_top)):,}")
if topups_saltados:
    print(f"       sin top-up por imports ANULADOS en la app (a revisión humana): {len(topups_saltados)} → {topups_saltados[:6]}")
print(f"  6) Finalizar (cerrados en Disapp con saldo acá): {len(finalizar)}  (saldo colgado ${round(sum(f['colgado'] for f in finalizar)):,})")
if no_finalizar_con_saldo:
    _colg = round(sum(x["colgado"] for x in no_finalizar_con_saldo))
    # La lista SIEMPRE se vuelca a CSV, se use o no el flag: es plata que alguien
    # debe y que un archivo no trae. Que exista el papel antes que la decisión.
    _csv = os.path.join(HERE, f"_ausentes_con_deuda_{SELLO}.csv")
    try:
        with open(_csv, "w", encoding="utf-8-sig", newline="") as fh:
            fh.write("prestamo_id,ref,colgado,cobrador_id,cliente_id\n")
            for x in sorted(no_finalizar_con_saldo, key=lambda y: -y["colgado"]):
                fh.write(f'{x["id"]},{x["ref"]},{x["colgado"]},{x["cobrador_id"]},{x["cliente_id"]}\n')
    except OSError as e:
        print(f"       (no pude escribir el CSV: {e})")
    if CERRAR_CON_SALDO:
        print(f"       ⚠️ --cerrar-con-saldo: SE CIERRAN {len(no_finalizar_con_saldo)} créditos que aún deben ${_colg:,}")
        print(f"          esa deuda se da por terminada. Lista completa → {_csv}")
    else:
        print(f"       🛡️ NO se cierran por tener DEUDA VIVA (candado 08-05): {len(no_finalizar_con_saldo)}  ${_colg:,}")
        print(f"          → revisar con Mauricio si Disapp los dio de baja o si hay que seguir cobrándolos")
        print(f"          lista completa → {_csv}")
        print(f"          si el export es COMPLETO y confiable (recarga total), re-correr con --cerrar-con-saldo")
print(f"       clientes borrados en Disapp que se MANTIENEN activos: {len(mantener_borrados)}")

# end-state esperado contra el export de hoy
exact = under = over = 0
for ref, c in ref2hoy.items():
    if ref not in by_ref_db and ref not in refs_crear:
        continue
    gap = round(c["pagos"] - sim_final.get(ref, round(float(by_ref_db[ref]["pagado_acum"] or 0), 2) if ref in by_ref_db else 0.0), 2)
    if abs(gap) <= TOL: exact += 1
    elif gap > 0: under += 1
    else: over += 1
n_post = sum(1 for _, p in insertar if p["fecha"] > CORTE)
print(f"\n  END-STATE esperado vs 'Pagos' Disapp (al corte {CORTE}): exactos {exact} · cortos {under} · pasados {over} (de {exact+under+over})")
if n_post:
    print(f"  (recaudos POSTERIORES al corte incluidos sin cap: {n_post} — contra un target viejo van a figurar como 'pasados': esperado)")

huerf = {p["ref"] for p in d["pagos"].values() if p["ref"] and p["fecha"] and p["fecha"] >= FRONTERA
         and p["ref"] not in by_ref_db and p["ref"] not in ref2hoy}
print(f"  (pendiente estructural: {len(huerf)} refs nacieron y cerraron en la ventana — necesitan export TODOS-los-estados)")

if choques:
    print(f"\n🔴 GUARDIA: {len(choques)} recaudos chocan crédito+día con pagos NATIVOS de la app.")
    for ref, p in choques[:8]:
        print(f"     {ref} {p['fecha']} ${round(p['monto'] or 0):,}")
    if SALTEAR:
        # En zona viva la APP es la verdad: se saltean SOLO las filas que chocan
        # (posibles re-capturas en Disapp) y el resto del día entra normal. Así
        # un único choque no frena el import de todas las zonas.
        ids_choque = {id(p) for _, p in choques}
        insertar = [(r, p) for r, p in insertar if id(p) not in ids_choque]
        print(f"   --saltear-choques: se saltean esas {len(choques)} filas (${round(sum(p['monto'] or 0 for _, p in choques)):,}) y sigue el resto.")
    elif not FORZAR:
        sys.exit("   Abortado. Opciones: --saltear-choques (saltea SOLO esas filas y sigue) o --forzar (importa TODO, a sabiendas).")
else:
    print(f"\n  ✓ guardia: sin choques con pagos nativos ({len(nativos)} revisados)")

if not COMMIT:
    print("\n🟡 DRY-RUN: no se escribió nada. Aplicar con --commit.")
    sys.exit(0)

# ══ APLICAR ═════════════════════════════════════════════════════════════════
print("\n═══ APLICANDO ═══")
# 1) clientes (upsert ignore por disapp_id) y re-fetch del mapa
if cli_nuevos:
    E.upsert(db, "clientes", list(cli_nuevos.values()), "disapp_id", ignore=True, rep=False)
    clientes_db = E.get_rows(db, "clientes", "id,disapp_id")
    cliDe = {str(c["disapp_id"]): c["id"] for c in clientes_db if c.get("disapp_id")}

# 2) créditos (insert plano; idempotencia por el pre-filtro contra la base)
filas_credito = []
for c in crear:
    cliente_id = cliDe.get(c["id_cliente"])
    if not cliente_id:
        print(f"  ✗ crédito {c['cid']}: cliente disapp {c['id_cliente']} no resuelto")
        continue
    filas_credito.append({
        "cliente_id": cliente_id,
        "cobrador_id": vend2user[str(c["id_vendedor"])]["id"],
        "monto_prestado": round(c["valor"]),
        "cuota_diaria": round(c["cuota"]),
        "total_dias": c["cuotas"],
        "fecha_inicio": (c["fecha"] or dt.date(2026, 7, 1)).isoformat(),
        "frecuencia": c["frecuencia"],
        "estado": "activo",
        "interes_pct": c["tasa"] if c["tasa"] is not None else 20,
        "disapp_credit_id": c["cid"],
        "disapp_credit_ref": c["ref"],
    })
creados = []
for i in range(0, len(filas_credito), 200):
    st, data = E.http(db["url"] + "/rest/v1/prestamos", "POST", key, filas_credito[i:i+200],
                      "return=representation")
    if st >= 300:
        raise RuntimeError(f"insert prestamos [{st}]: {str(data)[:400]}")
    creados.extend(data)
    print(f"    prestamos: {min(i+200, len(filas_credito))}/{len(filas_credito)}", end="\r")
print(f"\n  ✓ créditos creados: {len(creados)}")
for row in creados:
    if row.get("disapp_credit_ref"):
        by_ref_db[row["disapp_credit_ref"]] = row

# 2b) resucitar
okR = 0
for r in resucitar:
    st, _ = E.http(db["url"] + f"/rest/v1/prestamos?id=eq.{r['id']}", "PATCH", key,
                   {"estado": "activo", "finalizado_en": None}, "return=minimal")
    if st < 300:
        okR += 1
    else:
        print(f"  ✗ resucitar {r['ref']}: [{st}]")
print(f"  ✓ resucitados: {okR}/{len(resucitar)}")

# 3) asignaciones
ahora_iso = dt.datetime.now(UY).isoformat()
filas_asig = []
for a in asig_crear:
    cliente_id = a.get("cliente_id") or cliDe.get(a.get("cliente_disapp", ""))
    if not cliente_id or (cliente_id, a["cobrador_id"]) in act_pair:
        continue
    filas_asig.append({"cliente_id": cliente_id, "cobrador_id": a["cobrador_id"], "activo": True, "asignado_en": ahora_iso})
if filas_asig:
    st, data = E.http(db["url"] + "/rest/v1/asignaciones", "POST", key, filas_asig, "return=minimal")
    if st >= 300:
        raise RuntimeError(f"insert asignaciones [{st}]: {str(data)[:400]}")
for a in asig_activar:
    E.http(db["url"] + f"/rest/v1/asignaciones?id=eq.{a['id']}", "PATCH", key, {"activo": True}, "return=minimal")
for a in asig_bajar:
    E.http(db["url"] + f"/rest/v1/asignaciones?id=eq.{a['id']}", "PATCH", key, {"activo": False}, "return=minimal")
print(f"  ✓ asignaciones: +{len(filas_asig)} nuevas · {len(asig_activar)} reactivadas · {len(asig_bajar)} bajadas")

# 4) recaudos (upsert ignore por disapp_pago_id)
vend_txt2id = d["vend_por_texto"]
filas_pago = []
for ref, p in insertar:
    pdb = by_ref_db.get(ref)
    if not pdb:
        continue  # crédito que no se pudo crear
    td = int(pdb.get("total_dias") or 10**6)
    dc = p["cuota_num"] or 1
    dc = max(1, min(dc, td))
    idv = vend_txt2id.get(p["vendedor"])
    u = vend2user.get(str(idv)) if idv is not None else None
    filas_pago.append({
        "prestamo_id": pdb["id"],
        "dia_credito": dc,
        "monto": p["monto"] or 0.01,
        "registrado_por": (u or {}).get("id") or pdb.get("cobrador_id"),
        "registrado_en": E.iso_ts(p["fecha"]),
        "origen": "disapp_import",
        "importado_en": dt.datetime.now().isoformat(),
        "disapp_pago_id": p["id_pago"],
        "disapp_credit_ref": ref,
    })
antes = E.count_rows(db, "pagos")
E.upsert(db, "pagos", filas_pago, "disapp_pago_id", ignore=True, rep=False)
despues = E.count_rows(db, "pagos")
print(f"  ✓ recaudos: {despues - antes} insertados (candidatos {len(filas_pago)}; el resto ya existía)")

# 5) top-ups exactos
filas_top = []
for t in topups:
    pdb = by_ref_db.get(t["ref"])
    if not pdb:
        continue
    filas_top.append({
        "prestamo_id": pdb["id"],
        "dia_credito": 1,
        "monto": t["delta"],
        "registrado_por": pdb.get("cobrador_id"),
        # Fechado AL CORTE (el día cuyos libros reconcilia), nunca HOY: un
        # asiento fechado hoy entraba a la custodia del día (esperado del
        # cierre, float alto, push de las 21:00) como plata en mano fantasma.
        "registrado_en": E.iso_ts(CORTE),
        "origen": "reconciliacion_0804",
        "importado_en": dt.datetime.now().isoformat(),
        # El id lleva el DELTA: con solo recon-{día}-{cid}, la SEGUNDA corrida
        # del mismo día (créditos frescos a la noche tras el import de la
        # madrugada) chocaba con el ajuste de la primera y el upsert-ignore la
        # tiraba en silencio → 63 créditos quedaban "cortos" para siempre.
        # Mismo delta re-intentado sigue colisionando (idempotente, correcto).
        "disapp_pago_id": f"recon-{SELLO}-{t['cid']}-{int(round(t['delta'] * 100))}",
        "disapp_credit_ref": t["ref"],
    })
E.upsert(db, "pagos", filas_top, "disapp_pago_id", ignore=True, rep=False)
print(f"  ✓ top-ups: {len(filas_top)}")

# 6) finalizar
okF = 0
for f in finalizar:
    st, _ = E.http(db["url"] + f"/rest/v1/prestamos?id=eq.{f['id']}&estado=eq.activo", "PATCH", key,
                   {"estado": "finalizado", "finalizado_en": ahora_iso}, "return=minimal")
    if st < 300:
        okF += 1
print(f"  ✓ finalizados: {okF}/{len(finalizar)}")

# ══ VERIFICACIÓN end-state real ═════════════════════════════════════════════
print("\n═══ VERIFICACIÓN (contra la base, no simulada) ═══")
pres2 = E.get_rows(db, "prestamos", "id,cliente_id,disapp_credit_ref,estado,cuota_diaria,total_dias,pagado_acum,cobrador_id")
by_ref2 = {p["disapp_credit_ref"]: p for p in pres2 if p.get("disapp_credit_ref")}
exact = under = over = 0
peor = []
for ref, c in ref2hoy.items():
    p = by_ref2.get(ref)
    if not p:
        continue
    gap = round(c["pagos"] - float(p["pagado_acum"] or 0), 2)
    if abs(gap) <= TOL: exact += 1
    elif gap > 0: under += 1; peor.append((ref, gap))
    else: over += 1; peor.append((ref, gap))
print(f"  vs 'Pagos' Disapp hoy: EXACTOS {exact} · cortos {under} · pasados {over}")
for ref, g in sorted(peor, key=lambda x: -abs(x[1]))[:8]:
    print(f"    {ref}: {'falta' if g > 0 else 'sobra'} ${round(abs(g)):,}")
act2 = [p for p in pres2 if p["estado"] == "activo"]
cart = sum(max(0, round(float(p["cuota_diaria"] or 0)) * int(p["total_dias"] or 0) - round(float(p["pagado_acum"] or 0))) for p in act2)
print(f"  activos: {len(act2)} · cartera (cuota×días−pagado): ${cart:,}")

# SALUD DE FORMA (auditoría 08-05): multi-activo y sobre-cobro son las dos
# deformaciones que el empalme puede fabricar sin tocar un peso — se miden
# SIEMPRE y se comparan contra el export (la verdad de Disapp).
multi_cli = defaultdict(int)
for p in act2:
    multi_cli[p["cliente_id"]] += 1
n_multi = sum(1 for v in multi_cli.values() if v > 1)
multi_export = defaultdict(int)
for c in ref2hoy.values():
    # Por ID CLIENTE de Disapp (no "cid", que es el id del CRÉDITO y daba
    # siempre 1 por fila → "Disapp: 0" mentiroso en la primera corrida).
    multi_export[c["id_cliente"] or c["cid"]] += 1
n_multi_export = sum(1 for v in multi_export.values() if v > 1)
sobre2 = sum(
    1 for p in act2
    if float(p["pagado_acum"] or 0) > round(float(p["cuota_diaria"] or 0)) * int(p["total_dias"] or 0) + 0.5
)
print(f"  clientes con >1 crédito activo: {n_multi} (Disapp: {n_multi_export}) · activos sobre-cobrados: {sobre2}")
if n_multi > n_multi_export + 5 or sobre2 > 0:
    print("  ⚠️  DEFORMACIÓN: la base tiene más multi-activos que Disapp o hay sobre-cobros activos. Revisar antes de dar por buena la corrida.")

# LOG DE CORRIDA en disco (antes no quedaba rastro reproducible de qué se hizo).
import json as _json
_ahora = locals().get("ahora_iso") or dt.datetime.now().isoformat()
_log = {
    "corrida": _ahora, "corte": str(CORTE), "commit": COMMIT,
    "finalizados_plan": len(finalizar), "resucitados_plan": len(resucitar),
    "top_ups_plan": len(topups), "mantener_borrados": len(mantener_borrados),
    "activos_final": len(act2), "cartera": cart,
    "clientes_multi_activo": n_multi, "multi_activo_disapp": n_multi_export,
    "activos_sobre_cobrados": sobre2,
    "vs_pagos_disapp": {"exactos": exact, "cortos": under, "pasados": over},
}
_ruta_log = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"_empalme_log_{_ahora[:10]}.json")
with open(_ruta_log, "w", encoding="utf-8") as _f:
    _json.dump(_log, _f, ensure_ascii=False, indent=1)
print(f"  log de corrida: {_ruta_log}")
print("\nSiguiente: correr scripts/shadow-disapp.py --env-file .env.local --src <creditos de hoy> para el diff independiente.")
