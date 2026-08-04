#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RECONSTRUIR créditos "muertos" desde sus propios recaudos (sin pedirle nada a
Disapp). Son créditos que nacieron y se cerraron entre dos exports de créditos
(p.ej. 07-21→08-03): no figuran en el export de activos de HOY ni existen en la
base, así que sus recaudos (~$942k) no tenían dónde caer.

La jugada: cada fila de recaudo trae el crédito completo — 'Total Crédito',
'Valor Cuota', 'Total Cuotas', 'Cliente', 'Documento', 'Teléfono', 'Vendedor',
'Cuota #', 'Fecha Pago'. Con eso se reconstruye el crédito (FINALIZADO, con
fecha estimada por la cadencia de pagos) y se le imputan TODOS sus recaudos.
El cliente se matchea por DOCUMENTO (nunca por nombre); si no existe, se crea
con disapp_id sintético 'recon-<ref>' (idempotencia + trazabilidad).

Alcance: refs con actividad desde 2026-07-21 (los muertos de la transición).
Los finalizados antiguos sin actividad reciente no aportan nada operativo.

  Dry-run:  python scripts/reconstruir-creditos-muertos.py --env-file .env.local
  Escribir: ... --commit
"""
import sys, os, glob, datetime as dt, urllib.parse
from collections import defaultdict, Counter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import empalme_disapp as E
import openpyxl

def arg(flag, default=None):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else default

COMMIT = "--commit" in sys.argv
SRC = arg("--src", r"C:\Users\Carlos\migracion")
ENVF = arg("--env-file", ".env.local")
ACTIVIDAD_DESDE = dt.date(2026, 7, 21)

env = E.load_env(ENVF)
url = (env.get("SUPABASE_URL") or env.get("NEXT_PUBLIC_SUPABASE_URL") or "").rstrip("/")
key = env.get("SUPABASE_SERVICE_ROLE_KEY")
if not url or not key:
    sys.exit(f"Faltan SUPABASE_URL/SERVICE_ROLE_KEY en {ENVF}")
db = {"url": url, "key": key, "host": urllib.parse.urlparse(url).netloc}
print(f"RECONSTRUIR MUERTOS → {db['host'].split('.')[0]}  | modo: {'🔴 COMMIT' if COMMIT else '🟡 DRY-RUN'}")

# ── 1. Recaudos con datos completos (re-leo los xlsx: el parser común no trae
#       Cliente/Teléfono, acá hacen falta para poder crear al cliente) ────────
pagos = {}
for path in sorted(glob.glob(os.path.join(SRC, "recaudos*.xlsx"))):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdr = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    hm = {h.lower(): i for i, h in enumerate(hdr) if h}
    def gv(r, *names):
        for n in names:
            i = hm.get(n.lower())
            if i is not None:
                return r[i]
        return None
    for r in ws.iter_rows(min_row=2, values_only=True):
        pid = gv(r, "ID Pago")
        if pid in (None, ""):
            continue
        pid = str(pid).strip()
        if pid in pagos:
            continue
        pagos[pid] = {
            "id_pago": pid,
            "ref": (str(gv(r, "Ref. Crédito", "Ref. Credito") or "")).strip() or None,
            "vendedor": (str(gv(r, "Vendedor") or "")).strip(),
            "cliente": (str(gv(r, "Cliente") or "")).strip(),
            "documento": (str(gv(r, "Documento") or "")).strip() or None,
            "telefono": (str(gv(r, "Teléfono", "Telefono") or "")).strip() or None,
            "total_credito": E.parse_recaudo(gv(r, "Total Crédito", "Total Credito")),
            "valor_cuota": E.parse_recaudo(gv(r, "Valor Cuota")),
            "monto": E.parse_recaudo(gv(r, "Recaudo")),
            "cuota_num": int(gv(r, "Cuota #") or 0) or None,
            "total_cuotas": int(gv(r, "Total Cuotas") or 0) or None,
            "fecha": E.parse_date(gv(r, "Fecha Pago")),
        }
    wb.close()
print(f"  recaudos únicos en carpeta: {len(pagos)}")

# ── 2. Refs "muertas": ni en la base ni en el export de créditos de HOY ─────
refs_file_hoy = set()
for path in sorted(glob.glob(os.path.join(SRC, "creditos*.xlsx"))):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hdr = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    hm = {h.lower(): i for i, h in enumerate(hdr) if h}
    iref = hm.get("crédito #") or hm.get("credito #")
    for r in ws.iter_rows(min_row=2, values_only=True):
        if iref is not None and r[iref] not in (None, ""):
            refs_file_hoy.add(str(r[iref]).strip())
    wb.close()

pres = E.get_rows(db, "prestamos", "id,disapp_credit_ref")
refs_db = {p["disapp_credit_ref"] for p in pres if p.get("disapp_credit_ref")}

por_ref = defaultdict(list)
for p in pagos.values():
    if p["ref"] and p["ref"] not in refs_db and p["ref"] not in refs_file_hoy:
        por_ref[p["ref"]].append(p)
# solo los muertos DE LA TRANSICIÓN (actividad reciente)
por_ref = {ref: sorted(lst, key=lambda p: (p["fecha"] or dt.date.min, p["id_pago"]))
           for ref, lst in por_ref.items()
           if any(p["fecha"] and p["fecha"] >= ACTIVIDAD_DESDE for p in lst)}
print(f"  refs muertas con actividad ≥ {ACTIVIDAD_DESDE}: {len(por_ref)}")

# ── 3. Reconstruir cada crédito desde sus recaudos ──────────────────────────
usuarios = E.get_rows(db, "usuarios", "id,nombre,disapp_vendedor_id")
# vendedor por TEXTO → usuario: el texto de recaudos coincide con el de créditos;
# el mapa texto→id viene del consolidado del empalme (créditos de hoy).
d = E.consolidar(SRC)
vend_txt2id = d["vend_por_texto"]
vend2user = {str(u["disapp_vendedor_id"]): u for u in usuarios if u.get("disapp_vendedor_id") is not None}

clientes_db = E.get_rows(db, "clientes", "id,disapp_id,documento,nombre")
por_doc = {c["documento"]: c["id"] for c in clientes_db if c.get("documento")}
por_disapp = {str(c["disapp_id"]): c["id"] for c in clientes_db if c.get("disapp_id")}

def moda(vals):
    vals = [v for v in vals if v not in (None, 0, "")]
    return Counter(vals).most_common(1)[0][0] if vals else None

def cadencia(fechas):
    """diaria/semanal/quincenal/mensual según el gap mediano entre pagos."""
    fs = sorted({f for f in fechas if f})
    if len(fs) < 2:
        return "diario", 1
    gaps = sorted((b - a).days for a, b in zip(fs, fs[1:]) if (b - a).days > 0)
    g = gaps[len(gaps) // 2] if gaps else 1
    if g <= 3:
        return "diario", 1
    if g <= 10:
        return "semanal", 7
    if g <= 20:
        return "quincenal", 15
    return "mensual", 30

plan, sin_vendedor, inconsistentes = [], [], 0
for ref, lst in sorted(por_ref.items()):
    cuota = moda([p["valor_cuota"] for p in lst]) or 0
    cuotas = int(moda([p["total_cuotas"] for p in lst]) or 0)
    total = moda([p["total_credito"] for p in lst]) or (cuota * cuotas)
    if cuota and cuotas and total and abs(cuota * cuotas - total) > 1:
        inconsistentes += 1
    vend_txt = moda([p["vendedor"] for p in lst]) or ""
    u = vend2user.get(str(vend_txt2id.get(vend_txt, "")))
    if not u:
        sin_vendedor.append((ref, vend_txt))
        continue
    frec, paso = cadencia([p["fecha"] for p in lst])
    primera = min(p["fecha"] for p in lst if p["fecha"])
    ultima = max(p["fecha"] for p in lst if p["fecha"])
    primer_num = min((p["cuota_num"] or 1) for p in lst)
    fecha_inicio = primera - dt.timedelta(days=(max(primer_num, 1) - 1) * paso)
    doc = moda([p["documento"] for p in lst])
    cliente_id = por_doc.get(doc) if doc else None
    if not cliente_id:
        cliente_id = por_disapp.get(f"recon-{ref}")  # idempotencia si ya lo creamos
    suma = round(sum(p["monto"] or 0 for p in lst), 2)
    plan.append({
        "ref": ref, "lst": lst, "cuota": round(cuota), "cuotas": cuotas,
        "total": round(total), "monto": round(total / 1.2), "frec": frec,
        "fecha_inicio": fecha_inicio, "ultima": ultima, "cobrador": u["id"],
        "cliente_id": cliente_id, "doc": doc,
        "cliente_nombre": moda([p["cliente"] for p in lst]) or "(sin nombre)",
        "telefono": moda([p["telefono"] for p in lst]),
        "suma": suma, "cobertura": (suma / total if total else 0),
    })

suma_total = sum(p["suma"] for p in plan)
print(f"\n═══ PLAN ═══")
print(f"  créditos a reconstruir (FINALIZADOS): {len(plan)}  · recaudos a imputar: {sum(len(p['lst']) for p in plan)}  ${round(suma_total):,}")
print(f"  clientes: existentes por documento {sum(1 for p in plan if p['cliente_id'])} · a CREAR {sum(1 for p in plan if not p['cliente_id'])}")
print(f"  cuota×cuotas ≠ Total Crédito (informativo): {inconsistentes}")
print(f"  por cadencia: {dict(Counter(p['frec'] for p in plan))}")
cob_full = sum(1 for p in plan if p["cobertura"] >= 0.99)
print(f"  cobertura de recaudos vs total: pagados-completos {cob_full} · parciales {len(plan) - cob_full} (el resto se fue en renovación/cierre allá)")
if sin_vendedor:
    print(f"  ⚠ sin vendedor mapeado (NO se reconstruyen): {len(sin_vendedor)} → {sin_vendedor[:5]}")
for p in plan[:8]:
    print(f"    {p['ref']}: {p['frec']} ${p['cuota']:,}×{p['cuotas']} total ${p['total']:,} · pagó ${round(p['suma']):,} ({p['cobertura']:.0%}) · {p['fecha_inicio']}→{p['ultima']} · cliente {'OK' if p['cliente_id'] else 'CREAR'}")

if not COMMIT:
    print("\n🟡 DRY-RUN: nada escrito. Aplicar con --commit.")
    sys.exit(0)

# ── 4. Aplicar ──────────────────────────────────────────────────────────────
print("\n═══ APLICANDO ═══")
UY = dt.timezone(dt.timedelta(hours=-3))
docs_usados = set(por_doc.keys())
creados_cli = 0
for p in plan:
    if p["cliente_id"]:
        continue
    doc = p["doc"] if (p["doc"] and p["doc"] not in docs_usados) else None
    if doc:
        docs_usados.add(doc)
    st, data = E.http(db["url"] + "/rest/v1/clientes", "POST", key, {
        "nombre": p["cliente_nombre"], "documento": doc, "telefono": p["telefono"],
        "disapp_id": f"recon-{p['ref']}", "origen": "oficina", "activo": True,
    }, "return=representation")
    if st >= 300:
        print(f"  ✗ cliente p/ {p['ref']}: [{st}] {str(data)[:150]}")
        continue
    p["cliente_id"] = data[0]["id"]
    creados_cli += 1
print(f"  ✓ clientes creados: {creados_cli}")

okC = okP = 0
filas_pago = []
for p in plan:
    if not p["cliente_id"]:
        continue
    st, data = E.http(db["url"] + "/rest/v1/prestamos", "POST", key, {
        "cliente_id": p["cliente_id"], "cobrador_id": p["cobrador"],
        "monto_prestado": p["monto"], "cuota_diaria": p["cuota"],
        "total_dias": p["cuotas"], "fecha_inicio": p["fecha_inicio"].isoformat(),
        "frecuencia": p["frec"], "estado": "finalizado",
        "finalizado_en": f"{p['ultima'].isoformat()}T12:00:00-03:00",
        "interes_pct": 20,
        "disapp_credit_id": f"recon-{p['ref']}", "disapp_credit_ref": p["ref"],
    }, "return=representation")
    if st >= 300:
        print(f"  ✗ crédito {p['ref']}: [{st}] {str(data)[:180]}")
        continue
    okC += 1
    pid = data[0]["id"]
    for pg in p["lst"]:
        dc = pg["cuota_num"] or 1
        dc = max(1, min(dc, p["cuotas"] or 10**6))
        filas_pago.append({
            "prestamo_id": pid, "dia_credito": dc, "monto": pg["monto"] or 0.01,
            "registrado_por": p["cobrador"], "registrado_en": E.iso_ts(pg["fecha"]),
            "origen": "disapp_import", "importado_en": dt.datetime.now().isoformat(),
            "disapp_pago_id": pg["id_pago"], "disapp_credit_ref": p["ref"],
        })
print(f"  ✓ créditos reconstruidos: {okC}/{len(plan)}")
antes = E.count_rows(db, "pagos")
E.upsert(db, "pagos", filas_pago, "disapp_pago_id", ignore=True, rep=False)
print(f"  ✓ recaudos imputados: {E.count_rows(db, 'pagos') - antes} (candidatos {len(filas_pago)})")

# verificación: pagado_acum de los reconstruidos == suma de sus recaudos
pres2 = E.get_rows(db, "prestamos", "id,disapp_credit_ref,pagado_acum,estado")
mapa = {p["disapp_credit_ref"]: p for p in pres2 if (p.get("disapp_credit_ref") or "").strip()}
drift = 0
for p in plan:
    row = mapa.get(p["ref"])
    if row and abs(float(row["pagado_acum"] or 0) - p["suma"]) > 1:
        drift += 1
        print(f"  ✗ {p['ref']}: pagado_acum={row['pagado_acum']} vs Σrecaudos={p['suma']}")
print(f"\n  VERIFICACIÓN: drift={drift} (debe ser 0) · todos FINALIZADOS (no tocan rutas ni cartera activa)")
